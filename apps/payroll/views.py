import base64
import calendar
import io
import json
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

try:
    from num2words import num2words
except ImportError:
    num2words = None
import openpyxl
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import Q, Sum, Case, When, IntegerField, Value
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.http import JsonResponse
from django.http import Http404
from django.core import signing
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.html import strip_tags
from django.views.generic import ListView, TemplateView, View, DeleteView, UpdateView, CreateView, DetailView
from xhtml2pdf import pisa

from accounting.models import JournalItem
from budget.models import BudgetAssignmentHistory, BudgetGroup
from contract.models import ManagementPeriod
from core.models import CatalogItem
from employee.models import Employee
from .forms import PayrollPeriodForm, PayrollConstantForm, PayrollRubricForm
from .models import PayrollPeriod, Payslip, PayrollConstant, PayslipItem, PayrollNovelty, PayrollRubric
from .models import PendingDebt
from .services import PayrollCalculatorService
from .services import rebuild_accounting_for_period

PAYSLIP_PUBLIC_TOKEN_SALT = 'payroll.public.validation'


def build_public_payslip_token(payslip_id):
    """Genera token firmado para validacion publica del rol."""
    return signing.dumps({'payslip_id': payslip_id}, salt=PAYSLIP_PUBLIC_TOKEN_SALT)


def parse_public_payslip_token(token):
    """Obtiene el id del rol desde un token firmado."""
    payload = signing.loads(token, salt=PAYSLIP_PUBLIC_TOKEN_SALT)
    payslip_id = payload.get('payslip_id')
    if not payslip_id:
        raise signing.BadSignature('Token sin payslip_id')
    return int(payslip_id)


class PayrollListView(ListView):
    """Vista principal con renderizado híbrido"""
    model = Payslip
    template_name = 'payroll/payroll_list.html'
    context_object_name = 'payslips'
    paginate_by = 50

    def get_queryset(self):
        period_id = self.request.GET.get('period_id')
        if period_id:
            return Payslip.objects.filter(period_id=period_id).select_related('employee', 'period')
        return Payslip.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['periods'] = PayrollPeriod.objects.all()
        # Si hay periodo seleccionado, enviarlo al contexto
        if self.request.GET.get('period_id'):
            context['current_period'] = get_object_or_404(PayrollPeriod, pk=self.request.GET.get('period_id'))
        return context

    def render_to_response(self, context, **response_kwargs):
        """Híbrido: Si es AJAX devuelve solo la tabla, si no, la página entera"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('payroll/partial_payroll_table.html', context)
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)


class PeriodListView(ListView):
    model = PayrollPeriod
    template_name = 'payroll/payroll_period_list.html'
    context_object_name = 'periods'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = PayrollPeriodForm()
        for period in context['periods']:
            try:
                cutoff_date = date(int(period.year), period.month_number, 25)
                end_of_month = period.end_date
                has_changes = BudgetAssignmentHistory.objects.filter(
                    Q(start_date__gt=cutoff_date, start_date__lte=end_of_month) |
                    Q(end_date__gt=cutoff_date, end_date__lte=end_of_month)
                ).exists()
                period.has_scope_changes = has_changes
            except (ValueError, TypeError):
                period.has_scope_changes = False

        return context

    def get_queryset(self):
        qs = super().get_queryset()
        # Anotamos un número de mes para orden cronológica correcta (YYYYMM)
        month_case = Case(
            When(month='ENERO', then=Value(1)), When(month='FEBRERO', then=Value(2)),
            When(month='MARZO', then=Value(3)), When(month='ABRIL', then=Value(4)),
            When(month='MAYO', then=Value(5)), When(month='JUNIO', then=Value(6)),
            When(month='JULIO', then=Value(7)), When(month='AGOSTO', then=Value(8)),
            When(month='SEPTIEMBRE', then=Value(9)), When(month='OCTUBRE', then=Value(10)),
            When(month='NOVIEMBRE', then=Value(11)), When(month='DICIEMBRE', then=Value(12)),
            output_field=IntegerField()
        )
        qs = qs.annotate(month_num=month_case, year_int=Cast('year', IntegerField()))

        # Siempre retornamos todos los periodos ordenados
        return qs.order_by('-year_int', '-month_num')

    def get(self, request, *args, **kwargs):
        # Si es petición AJAX devolvemos el partial completo (HTML) empaquetado en JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # configurar object_list y contexto para la paginación de ListView
            self.object_list = self.get_queryset()
            context = self.get_context_data()

            # Si el cliente solicita JSON de datos (modo JS-render), devolvemos estructuras JSON
            if request.GET.get('json') == '1':
                periods = context.get('periods') or self.object_list
                data = []
                for p in periods:
                    month_num = getattr(p, 'month_num', None) or getattr(p, 'month_number', 0)
                    data.append({
                        'id': p.id,
                        'month': p.month,
                        'year': p.year,
                        'month_num': int(month_num),
                        'start_date': p.start_date.strftime('%d/%m/%Y') if p.start_date else '',
                        'end_date': p.end_date.strftime('%d/%m/%Y') if p.end_date else '',
                        'working_days': p.working_days,
                        'is_closed': bool(p.is_closed),
                        'display': f"{p.month} {p.year}",
                        'payslip_url': reverse('payroll:payslip_list') + f"?period_id={p.id}",
                        'novelty_url': reverse('payroll:novelty_mass_load') + f"?period_id={p.id}"
                    })
                return JsonResponse({'periods': data})

            # Renderizamos el partial completo con contexto
            html = render_to_string('payroll/partials/partial_period_table.html', context, request=request)
            return JsonResponse({'html': html})
        return super().get(request, *args, **kwargs)


class PeriodCreateView(View):
    template_name = 'payroll/modals/modal_period_form.html'

    def get(self, request, *args, **kwargs):
        """Devuelve el formulario del modal (usado por el fetch GET en JS)."""
        form = PayrollPeriodForm()
        html = render_to_string(self.template_name, {'form': form}, request=request)
        return HttpResponse(html)

    def post(self, request):
        form = PayrollPeriodForm(request.POST)
        if form.is_valid():
            period = form.save(commit=False)
            period.created_by = request.user.username
            period.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Periodo creado correctamente',
                'period': {
                    'id': period.id,
                    'str': str(period)
                }
            })
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


class GeneratePayrollView(View):
    def post(self, request):
        period_id = request.POST.get('period_id')
        is_scope_run = request.POST.get('is_scope_run', 'false').lower() == 'true'

        try:
            period = PayrollPeriod.objects.get(pk=period_id)
            if period.is_closed:
                return JsonResponse({'status': 'error', 'message': 'El periodo está cerrado.'}, status=400)
            valid_history_emp_ids = BudgetAssignmentHistory.objects.filter(
                start_date__lte=period.end_date
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=period.start_date)
            ).values_list('employee_id', flat=True)
            # Cargar empleados (activos o inactivos) que tuvieron partida en el mes
            employees = Employee.objects.filter(
                id__in=valid_history_emp_ids
            ).select_related('person', 'person__economic_data', 'person__economic_data__payroll_info')

            service = PayrollCalculatorService(period, employees, is_scope_run=is_scope_run)
            # Vuelve el resultado del servicio
            result = service.generate_bulk()

            warnings = result.get('warnings', [])
            msg = 'Cálculo completado exitosamente.'
            if is_scope_run:
                msg = 'Cálculo de alcance completado exitosamente.'

            if warnings:
                msg += ' (Se generaron advertencias contables, revisa los reportes)'

            return JsonResponse({
                'status': 'success',
                'message': msg,
                'warnings': warnings  # Enviamos esto al frontend para pintarlo en un modal o toast amarillo
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


class RubricListView(ListView):
    model = PayrollRubric
    template_name = 'payroll/rubric_list.html'
    context_object_name = 'rubrics'

    def get_queryset(self):
        qs = super().get_queryset()
        tipo = self.request.GET.get('tipo')  # Para que puedas filtrar en la tabla si quieres
        if tipo:
            qs = qs.filter(rubric_type=tipo)
        return qs


class RubricCreateView(CreateView):
    model = PayrollRubric
    form_class = PayrollRubricForm
    template_name = 'payroll/modals/modal_rubric_form.html'

    def form_valid(self, form):
        self.object = form.save()
        return JsonResponse({'status': 'success', 'message': 'Rubro creado correctamente.'})


class RubricUpdateView(UpdateView):
    model = PayrollRubric
    form_class = PayrollRubricForm
    template_name = 'payroll/modals/modal_rubric_form.html'

    def form_valid(self, form):
        self.object = form.save()
        return JsonResponse({'status': 'success', 'message': 'Rubro actualizado correctamente.'})


class PayrollGenerateForm(forms.Form):
    period_id = forms.IntegerField()


class GeneratePayrollUIView(View):
    """Interfaz simple para seleccionar empleados elegibles y generar rol con prorrateo."""
    template_name = 'payroll/generate_ui.html'

    def get(self, request):
        period_id = request.GET.get('period_id')
        periods = PayrollPeriod.objects.all()
        context = {'periods': periods}
        if period_id:
            period = PayrollPeriod.objects.get(pk=period_id)

            valid_emp_ids = BudgetAssignmentHistory.objects.filter(
                start_date__lte=period.end_date
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=period.start_date)
            ).values_list('employee_id', flat=True).distinct()

            employees = Employee.objects.filter(id__in=valid_emp_ids, is_active=True, person__is_active=True)

            rows = []
            for emp in employees:
                # determinar fecha de ingreso: usar el último ManagementPeriod firmado/activo
                join = None
                last_period = ManagementPeriod.objects.filter(employee=emp,
                                                              status__code__in=['FIRMADO', 'ACTIVO']).order_by(
                    '-start_date').first()
                if last_period and last_period.signed_document:
                    join = last_period.start_date
                else:
                    # fallback a campos en Employee o InstitutionalData
                    join = emp.date_joined
                    if not join:
                        inst = getattr(emp, 'institutional_data', None)
                        join = getattr(inst, 'entry_date', None) if inst else None
                # si no hay fecha, considerar completo
                if not join or join <= period.start_date:
                    worked = period.working_days  # Generalmente 30
                elif join > period.end_date:
                    continue
                else:
                    # Aplicamos la misma lógica comercial
                    if period.end_date.month == 2 and join.month == 2 and join.day >= 28:
                        worked = (30 - join.day) + 1
                    elif join.day == 31:
                        worked = 1
                    else:
                        worked = (30 - join.day) + 1

                rows.append({'employee': emp, 'worked_days': max(0, worked)})

            context.update({'current_period': period, 'rows': rows})

        return render(request, self.template_name, context)


class GeneratePayrollSelectedView(View):
    def post(self, request):
        period_id = request.POST.get('period_id')
        is_scope_run = request.POST.get('is_scope_run', 'false').lower() == 'true'
        selected = request.POST.getlist('employee')
        worked_map = {}
        for k, v in request.POST.items():
            if k.startswith('worked_'):
                emp_id = k.split('_', 1)[1]
                try:
                    worked_map[int(emp_id)] = int(v)
                except:
                    pass

        if not period_id:
            messages.error(request, 'Periodo no definido')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        period = PayrollPeriod.objects.get(pk=period_id)
        employees_with_days = []
        for emp_id in selected:
            try:
                emp = Employee.objects.get(pk=emp_id)
            except Employee.DoesNotExist:
                continue
            wd = worked_map.get(emp.id, period.working_days)
            employees_with_days.append((emp, wd))

        svc = PayrollCalculatorService(period, [e for e, d in employees_with_days], is_scope_run=is_scope_run)
        svc.generate_for_selected(employees_with_days)
        messages.success(request, 'Rol generado para empleados seleccionados.')
        return redirect(request.META.get('HTTP_REFERER', '/'))


class ConstantListView(ListView):
    model = PayrollConstant
    template_name = 'payroll/constant_list.html'
    context_object_name = 'constants'

    def get_queryset(self):
        qs = super().get_queryset().order_by('name')
        show_inactive = self.request.GET.get('show_inactive')

        try:
            if show_inactive and str(show_inactive).lower() in ['true', '1', 'on']:
                return qs.all()
            return qs.filter(is_active=True)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning('Error al filtrar PayrollConstant.is_active: %s', e)
            return qs

    def get(self, request, *args, **kwargs):
        # Si es petición AJAX devolvemos sólo las filas (partial)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            constants = self.get_queryset()
            html = render_to_string('payroll/partials/_constant_rows.html', {'constants': constants})
            from django.http import HttpResponse
            return HttpResponse(html)
        return super().get(request, *args, **kwargs)


class ConstantCreateView(CreateView):
    model = PayrollConstant
    form_class = PayrollConstantForm
    template_name = 'payroll/modals/modal_constant_form.html'

    def form_valid(self, form):
        self.object = form.save()
        return JsonResponse({'status': 'success', 'message': 'Constante creada correctamente.'})

    def form_invalid(self, form):
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


class ConstantUpdateView(UpdateView):
    model = PayrollConstant
    form_class = PayrollConstantForm
    template_name = 'payroll/modals/modal_constant_form.html'

    def form_valid(self, form):
        self.object = form.save()
        return JsonResponse({'status': 'success', 'message': 'Constante actualizada correctamente.'})

    def form_invalid(self, form):
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


class ConstantDeleteView(DeleteView):
    model = PayrollConstant

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            return JsonResponse({'status': 'success', 'message': 'Constante eliminada.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


class PayslipListView(LoginRequiredMixin, ListView):
    model = Payslip
    template_name = 'payroll/payslip_list.html'
    context_object_name = 'payslips'
    paginate_by = 15

    def get_paginate_by(self, queryset):
        """Allow returning all results when frontend requests full dataset (full=1)."""
        full = (self.request.GET.get('full') or '').lower()
        if full in ['1', 'true', 'yes']:
            return None
        return self.paginate_by

    def get_queryset(self):
        period_id = self.request.GET.get('period_id')
        if not period_id or period_id == "None":
            return Payslip.objects.none()

        queryset = Payslip.objects.filter(period_id=period_id).select_related(
            'employee__person', 'period'
        ).order_by('employee__person__last_name')

        # Filtro para roles de alcance
        is_scope_view = self.request.GET.get('scope', 'false').lower() == 'true'
        if is_scope_view:
            period = get_object_or_404(PayrollPeriod, pk=period_id)
            cutoff_date = date(int(period.year), period.month_number, 25)
            end_of_month = period.end_date

            # Empleados con cambios después del corte
            employees_with_changes = BudgetAssignmentHistory.objects.filter(
                Q(start_date__gt=cutoff_date, start_date__lte=end_of_month) |
                Q(end_date__gt=cutoff_date, end_date__lte=end_of_month)
            ).values_list('employee_id', flat=True).distinct()

            queryset = queryset.filter(employee_id__in=employees_with_changes)

        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(employee__person__first_name__icontains=q) |
                Q(employee__person__last_name__icontains=q) |
                Q(employee__person__document_number__icontains=q) |
                Q(items__budget_line__budget_group__short_code__icontains=q)
            ).distinct()
        regime = self.request.GET.get('regime', '').strip()
        if regime:
            queryset = queryset.filter(items__budget_line__regime_item_id=regime).distinct()
        show_withheld = (self.request.GET.get('show_withheld') or '').lower()
        if show_withheld in ['1', 'true', 'on', 'only']:
            queryset = queryset.filter(is_withheld=True)
        else:
            # Por defecto mostramos TODOS excepto los retenidos
            queryset = queryset.filter(is_withheld=False)

        # Soporte de ordenamiento desde el cliente (TableManager envía sort_field & sort_dir)
        sort_field = self.request.GET.get('sort_field') or self.request.GET.get('sort')
        sort_dir = (self.request.GET.get('sort_dir') or 'asc').lower()
        allowed = {'employee__person__last_name', 'employee__person__document_number',
                   'items__budget_line__budget_group__short_code', 'total_income', 'total_deduction', 'net_pay'}
        if sort_field in allowed:
            order = sort_field if sort_dir == 'asc' else f'-{sort_field}'
            try:
                queryset = queryset.order_by(order)
            except Exception:
                pass

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['budget_groups'] = BudgetGroup.objects.all()
        context['labor_regimes'] = CatalogItem.objects.filter(catalog__code='LABOR_REGIMES', is_active=True).order_by(
            'name')
        period_id = self.request.GET.get('period_id')
        if period_id and period_id != "None":
            context['current_period'] = PayrollPeriod.objects.filter(id=period_id).first()
            context['period_id'] = period_id
            # Pasamos la búsqueda para que el input no se borre
            context['search_query'] = self.request.GET.get('q', '')

        # Agregar totales para mostrar en la cabecera (evita usar filtros inexistentes en la plantilla)
        try:
            qs = self.get_queryset()
            context['total_roles'] = qs.count()
            total_liquidado = qs.aggregate(total=Sum('net_pay'))['total'] or 0
            context['total_liquidado'] = total_liquidado
        except Exception:
            context['total_roles'] = 0
            context['total_liquidado'] = 0

        return context

    # ---> ESTA ES LA PIEZA MÁGICA QUE FALTABA <---
    def render_to_response(self, context, **response_kwargs):
        """Si la petición es AJAX, devuelve solo la tabla en formato JSON"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Renderiza solo el HTML del pedacito de la tabla
            html = render_to_string('payroll/partials/partial_payslip_table.html', context, request=self.request)
            # Lo empaqueta en JSON para que tu JavaScript no explote
            return JsonResponse({
                'html': html,
                'total_roles': context.get('total_roles', 0),
                'total_liquidado': context.get('total_liquidado', 0)
            })

        # Si es una carga normal de la página, hace lo de siempre
        return super().render_to_response(context, **response_kwargs)


class PayslipDetailView(DetailView):
    model = Payslip
    template_name = 'payroll/modals/modal_payslip_detail.html'
    context_object_name = 'payslip'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incomes'] = self.object.items.filter(
            item_type='INCOME'
        ).order_by('rubric__order')

        context['deductions'] = self.object.items.filter(
            item_type='DEDUCTION'
        ).exclude(
            rubric__code__icontains='PATRONAL'
        ).order_by('rubric__order')

        return context


class FondosReservaListView(LoginRequiredMixin, ListView):
    """Lista de empleados activos mostrando Fondos de Reserva y Mensualiza Décimos."""
    model = Employee
    template_name = 'payroll/reserve_funds_list.html'
    context_object_name = 'employees'
    paginate_by = 10
    partial_template_name = 'payroll/partials/partial_reserve_funds_table.html'

    def get_queryset(self):
        qs = Employee.objects.filter(is_active=True, person__is_active=True)
        # Traer relaciones comúnmente usadas para evitar N+1
        qs = qs.select_related('person__economic_data__payroll_info', 'area').prefetch_related(
            'current_budget_line__position_item')
        # Soporte de búsqueda simple desde frontend
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(person__first_name__icontains=q) |
                Q(person__last_name__icontains=q) |
                Q(person__document_number__icontains=q)
            )
        return qs.order_by('person__last_name', 'person__first_name')

    def render_to_response(self, context, **response_kwargs):
        # Si es petición AJAX devolvemos solo el partial con la tabla y datos de paginación
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string(self.partial_template_name, context, request=self.request)
            page_obj = context.get('page_obj')
            if page_obj:
                pagination_data = {
                    'start_index': page_obj.start_index(),
                    'end_index': page_obj.end_index(),
                    'total_count': page_obj.paginator.count,
                    'current_page': page_obj.number,
                    'total_pages': page_obj.paginator.num_pages,
                    'has_previous': page_obj.has_previous(),
                    'has_next': page_obj.has_next(),
                }
            else:
                pagination_data = {
                    'start_index': 0,
                    'end_index': 0,
                    'total_count': 0,
                    'current_page': 1,
                    'total_pages': 1,
                    'has_previous': False,
                    'has_next': False,
                }

            return JsonResponse({'html': html, 'pagination': pagination_data})
        return super().render_to_response(context, **response_kwargs)

    def get_paginate_by(self, queryset):
        """Return None to disable server-side pagination and return all employees."""
        # Usar el valor definido en `paginate_by` para habilitar paginación servidor
        return self.paginate_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Fondos de Reserva'
        return context


class InstitutionalReportView(TemplateView):
    template_name = 'payroll/reports/institutional_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        period_id = self.kwargs.get('period_id')
        period = get_object_or_404(PayrollPeriod, pk=period_id)

        # 1. JORNALIZACIÓN
        jornalizacion_items = JournalItem.objects.filter(reference=str(period))
        jornalizacion = []
        total_debe = 0
        total_haber = 0

        if jornalizacion_items.exists():
            jornalizacion = jornalizacion_items.values(
                'account__code', 'account__name'
            ).annotate(total_debe=Sum('debit'), total_haber=Sum('credit')).order_by('account__code')
            total_debe = sum(item['total_debe'] for item in jornalizacion)
            total_haber = sum(item['total_haber'] for item in jornalizacion)

            # 2. DETALLE PRESUPUESTACIÓN (Mucha más rápida)
            presupuesto_list = []

            rubros_presupuestarios = PayslipItem.objects.filter(
                payslip__period=period,
                budget_line_code__isnull=False
            ).filter(
                Q(rubric__has_mapping=True) | Q(rubric__code__iexact='REMUNERACION')
            ).values(
                'budget_line_code', 'rubric__name'
            ).annotate(total=Sum('value'))

            for item in rubros_presupuestarios:
                presupuesto_list.append({
                    'partida': item['budget_line_code'],
                    'concepto': item['rubric__name'],
                    'monto': item['total']
                })

            presupuesto_list = sorted(presupuesto_list, key=lambda x: x['partida'])

        context.update({
            'period': period,
            'jornalizacion': jornalizacion,
            'total_debe': total_debe,
            'total_haber': total_haber,
            'presupuestacion': presupuesto_list,
        })
        return context


class NoveltyMassLoadView(TemplateView):
    template_name = 'payroll/novelty_mass_load.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['periods'] = PayrollPeriod.objects.filter(is_closed=False)

        context['incomes'] = PayrollRubric.objects.filter(
            rubric_type='INCOME',
            is_active=True,
            is_upload=True
        ).order_by('name')

        context['deductions'] = PayrollRubric.objects.filter(
            rubric_type='DEDUCTION',
            is_active=True,
            is_upload=True
        ).order_by('name')

        context['selected_period_id'] = self.request.GET.get('period_id', '')
        return context


class ParseNoveltyExcelView(View):
    def post(self, request):
        excel_file = request.FILES.get('file')
        rubro_type = request.POST.get('rubro_type')
        rubro_id = request.POST.get('rubro_id')
        if not excel_file: return JsonResponse({'status': 'error', 'message': 'Sin archivo'})

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            data_map, not_found = {}, set()

            # 1. Identificar modo de horas
            is_overtime_mode = False
            if rubro_type == 'INCOME' and rubro_id:
                rubric = PayrollRubric.objects.filter(pk=rubro_id).first()
                if rubric and getattr(rubric, 'is_overtime', False):
                    is_overtime_mode = True

            def to_decimal(value):
                try:
                    return Decimal(str(value or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                except:
                    return Decimal('0.00')

            # ==========================================
            # 🚀 FASE 1: LECTURA Y EXTRACCIÓN MASIVA (BULK)
            # ==========================================
            excel_rows = []
            cedulas_set = set()

            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not row[0]: continue
                raw_cedula = str(row[0]).strip()
                if not raw_cedula.replace('.', '').isdigit(): continue
                if raw_cedula.endswith('.0'): raw_cedula = raw_cedula[:-2]

                cedula = raw_cedula.zfill(10)
                cedulas_set.add(cedula)

                excel_rows.append({
                    'cedula': cedula,
                    'val_b': to_decimal(row[1] if len(row) > 1 else 0),
                    'val_c': to_decimal(row[2] if len(row) > 2 else 0)
                })

            # ==========================================
            # 🚀 FASE 2: PRECARGA EN MEMORIA (DICCIONARIOS)
            # ==========================================

            # A. Buscamos TODOS los empleados de una sola vez
            employees_qs = Employee.objects.select_related('person').filter(
                person__document_number__in=cedulas_set
            ).order_by('-is_active')

            emp_dict = {}
            for e in employees_qs:
                doc = e.person.document_number
                if doc not in emp_dict:
                    emp_dict[doc] = e

            # B. Buscamos TODAS las partidas de una sola vez (si es modo horas)
            bah_dict = {}
            if is_overtime_mode and emp_dict:
                from budget.models import BudgetAssignmentHistory
                emp_ids = [e.id for e in emp_dict.values()]

                # select_related para traer la info de la partida sin hacer más consultas
                bahs_qs = BudgetAssignmentHistory.objects.select_related('budget_line').filter(
                    employee_id__in=emp_ids
                ).order_by('employee_id', '-start_date')

                for bah in bahs_qs:
                    if bah.employee_id not in bah_dict:
                        bah_dict[bah.employee_id] = bah  # Guardamos solo la más reciente por empleado

            # ==========================================
            # 🚀 FASE 3: PROCESAMIENTO EN RAM (CERO CONSULTAS DB)
            # ==========================================
            for row_data in excel_rows:
                cedula = row_data['cedula']
                val_b = row_data['val_b']
                val_c = row_data['val_c']

                # Buscamos en nuestro diccionario de memoria (¡Instántaneo!)
                emp = emp_dict.get(cedula)

                if emp:
                    valor_final = Decimal('0.00')

                    if is_overtime_mode:
                        # Obtenemos la partida desde nuestro diccionario en memoria
                        bah = bah_dict.get(emp.id)

                        base_salary = Decimal('460.00')
                        partida_code = ""

                        if bah and bah.budget_line:
                            partida_code = bah.budget_line.code or ""
                            if bah.budget_line.remuneration:
                                base_salary = Decimal(str(bah.budget_line.remuneration))

                        # Matemática en memoria
                        hourly_value = (base_salary / Decimal('30.0') / Decimal('8.0')).quantize(Decimal('0.01'),
                                                                                                 rounding=ROUND_HALF_UP)
                        factor_extra = Decimal('1.60')
                        factor_supl = Decimal('2.00')

                        if "01.06" in partida_code:
                            factor_extra = Decimal('1.50')

                        total_dollars = Decimal('0.00')
                        if val_b > Decimal('0'):
                            total_dollars += val_b * hourly_value * factor_extra
                        if val_c > Decimal('0'):
                            total_dollars += val_c * hourly_value * factor_supl

                        valor_final = total_dollars.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                    else:
                        valor_final = val_b

                    # Sumamos al mapa de respuesta
                    if emp.id not in data_map:
                        data_map[emp.id] = {
                            'emp_id': emp.id,
                            'cedula': cedula,
                            'nombres': f"{emp.person.last_name} {emp.person.first_name}",
                            'valor': Decimal('0.00')
                        }

                    data_map[emp.id]['valor'] += valor_final
                else:
                    not_found.add(cedula)

            data = sorted([dict(item, valor=float(item['valor'])) for item in data_map.values()],
                          key=lambda x: x['nombres'])

            return JsonResponse({'status': 'success', 'data': data, 'not_found': sorted(list(not_found))})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class GetNoveltiesView(View):
    def get(self, request):
        period_id = request.GET.get('period_id')
        rubro_id = request.GET.get('rubro_id')
        if not all([period_id, rubro_id]): return JsonResponse({'status': 'error', 'message': 'Faltan parámetros'})

        try:
            novelties = PayrollNovelty.objects.filter(period_id=period_id, rubric_id=rubro_id,
                                                      value__gt=0).select_related('employee__person')
            data = sorted([{
                'emp_id': nov.employee.id, 'cedula': nov.employee.person.document_number,
                'nombres': f"{nov.employee.person.last_name} {nov.employee.person.first_name}",
                'valor': float(nov.value)
            } for nov in novelties], key=lambda x: x['nombres'])
            return JsonResponse({'status': 'success', 'data': data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class SaveNoveltiesView(View):
    def post(self, request):
        try:
            payload = json.loads(request.body)
            period_id = payload.get('period_id')
            rubro_id = payload.get('rubro_id')
            items = payload.get('items', [])

            period = PayrollPeriod.objects.get(pk=period_id)
            with transaction.atomic():
                PayrollNovelty.objects.filter(period=period, rubric_id=rubro_id).delete()
                novelties_to_create = []
                for item in items:
                    val, emp_id = Decimal(str(item.get('valor', 0))), item.get('emp_id')
                    if val > Decimal('0.00'):
                        novelties_to_create.append(
                            PayrollNovelty(period=period, employee_id=emp_id, rubric_id=rubro_id, value=val))
                PayrollNovelty.objects.bulk_create(novelties_to_create)
            return JsonResponse({'status': 'success', 'message': 'Novedades guardadas exitosamente'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class GroupedPayrollReportView(LoginRequiredMixin, View):
    def get(self, request, pk):
        period = get_object_or_404(PayrollPeriod, pk=pk)
        search_query = request.GET.get('q', '').strip()
        group_filter = request.GET.get('group', '').strip()
        regime_filter = request.GET.get('regime', '').strip()
        filter_type = request.GET.get('filtro', 'NORMAL')

        payslips_qs = Payslip.objects.filter(period=period)

        if search_query:
            payslips_qs = payslips_qs.filter(
                Q(employee__person__first_name__icontains=search_query) |
                Q(employee__person__last_name__icontains=search_query) |
                Q(employee__person__document_number__icontains=search_query) |
                Q(items__budget_line__budget_group__short_code__icontains=search_query)
            ).distinct()

        if group_filter:
            payslips_qs = payslips_qs.filter(items__budget_line__budget_group__short_code=group_filter).distinct()
        if regime_filter:
            payslips_qs = payslips_qs.filter(items__budget_line__regime_item_id=regime_filter).distinct()

        show_withheld = (self.request.GET.get('show_withheld') or '').lower()
        if show_withheld in ['only', '1', 'true', 'yes']:
            payslips_qs = payslips_qs.filter(is_withheld=True)
        else:
            if filter_type == 'NORMAL':
                payslips_qs = payslips_qs.filter(is_withheld=False)
            else:
                payslips_qs = payslips_qs.filter(is_withheld=False, is_paid=False)

        valid_payslip_ids = list(payslips_qs.values_list('id', flat=True))

        items_qs = PayslipItem.objects.filter(payslip_id__in=valid_payslip_ids).select_related(
            'payslip__employee__person', 'budget_line__budget_group', 'budget_line__spending_type_item',
            'rubric__debit_account', 'rubric__credit_account',
            'rubric__debit_account_prod', 'rubric__credit_account_prod',
            'rubric__debit_account_inv', 'rubric__credit_account_inv',
            'rubric__income_account'
        ).prefetch_related(
            'payslip__employee__person__economic_data__bank_account__bank',
            'payslip__employee__person__economic_data__bank_account__account_type'
        ).distinct()

        report_data = {}
        salary_rubrics = list(PayrollRubric.objects.filter(is_salary=True, is_active=True))

        # Función auxiliar interna para encontrar el rubro de sueldo según el tipo de gasto
        def _find_salary_rubric(spending_code: str):
            sc = str(spending_code or '5')
            if sc.startswith('7'):
                target = '7.1'
            elif sc.startswith('6'):
                target = '6.1'
            else:
                target = '5.1'

            return (
                    next((r for r in salary_rubrics if r.spending_context == target), None)
                    or next((r for r in salary_rubrics if r.spending_context == 'TODOS'), None)
            )

        # Procesamiento y agrupación de los rubros individuales de cada empleado
        for item in items_qs:
            val = Decimal(str(item.value))
            budget_group_obj = item.budget_line.budget_group if item.budget_line else None
            group_key = budget_group_obj.short_code if budget_group_obj else 'SIN_AGRUPAR'

            if group_key not in report_data:
                report_data[group_key] = {
                    'group_obj': budget_group_obj,
                    'group_name': budget_group_obj.name if budget_group_obj else 'Sin agrupación',
                    'empleados': {}, 'contabilidad': {}, 'presupuesto': {},
                    'ingresos_headers': {}, 'descuentos_headers': {}, 'aportes_headers': {}
                }

            group_data = report_data[group_key]
            employee_id = item.payslip.employee.id

            if employee_id not in group_data['empleados']:
                bank_name, account_type_name, account_number = "NO REGISTRADO", "", ""
                try:
                    bank_acc = item.payslip.employee.person.economic_data.bank_account
                    bank_name = bank_acc.bank.name if bank_acc.bank else "Desconocido"
                    account_type_name = bank_acc.account_type.name if bank_acc.account_type else ""
                    account_number = bank_acc.account_number
                except Exception:
                    pass

                group_data['empleados'][employee_id] = {
                    'persona': item.payslip.employee.person,
                    'empleado': item.payslip.employee,
                    'banco': bank_name, 'tipo_cuenta': account_type_name, 'numero_cuenta': account_number,
                    'ingresos': Decimal(0), 'descuentos': Decimal(0), 'liquido': Decimal(0),
                    'ingresos_dict': {}, 'descuentos_dict': {}, 'aportes_dict': {},
                    'net_pay_by_spending_type': {}  # Guarda el líquido separado por tipo de gasto (5.1, 7.1)
                }

            employee_dict = group_data['empleados'][employee_id]
            rubric_ref = item.rubric

            # Identificación del tipo de gasto (Corriente 5.1, Producción 6.1, Inversión 7.1)
            spending_type = '5.1'
            if item.budget_line and item.budget_line.spending_type_item:
                spending_type = item.budget_line.spending_type_item.code
            if rubric_ref.is_salary or 'main_spending_type' not in employee_dict:
                employee_dict['main_spending_type'] = spending_type
            # A. Distribución estructural en el Rol de Pagos (Sábana)
            if item.item_type == 'INCOME':
                header_key = f"INC_{rubric_ref.id}"
                order_val = rubric_ref.order if rubric_ref.order is not None else 999
                group_data['ingresos_headers'][header_key] = {
                    'key': header_key, 'name': rubric_ref.name, 'abbreviation': rubric_ref.abbreviation,
                    'order': order_val
                }
                employee_dict['ingresos_dict'][header_key] = employee_dict['ingresos_dict'].get(header_key,
                                                                                                Decimal(0)) + val
                employee_dict['ingresos'] += val
                employee_dict['liquido'] += val
                employee_dict['net_pay_by_spending_type'][spending_type] = employee_dict[
                                                                               'net_pay_by_spending_type'].get(
                    spending_type, Decimal(0)) + val

            elif item.item_type == 'DEDUCTION':
                header_key = f"DED_{rubric_ref.id}"
                order_val = rubric_ref.order if rubric_ref.order is not None else 999
                code_upper = (rubric_ref.code or '').upper()

                if 'IESS_PER' in code_upper or ('APORTE' in code_upper and 'PATRONAL' not in code_upper):
                    group_data['aportes_headers'][header_key] = {
                        'key': header_key, 'name': rubric_ref.name, 'order': order_val,
                        'abbreviation': rubric_ref.abbreviation
                    }
                    employee_dict['aportes_dict'][header_key] = employee_dict['aportes_dict'].get(header_key,
                                                                                                  Decimal(0)) + val
                else:
                    group_data['descuentos_headers'][header_key] = {
                        'key': header_key, 'name': rubric_ref.name, 'order': order_val,
                        'abbreviation': rubric_ref.abbreviation
                    }
                    employee_dict['descuentos_dict'][header_key] = employee_dict['descuentos_dict'].get(header_key,
                                                                                                        Decimal(
                                                                                                            0)) + val

                employee_dict['descuentos'] += val
                employee_dict['liquido'] -= val
                employee_dict['net_pay_by_spending_type'][spending_type] = employee_dict[
                                                                               'net_pay_by_spending_type'].get(
                    spending_type, Decimal(0)) - val

            elif item.item_type == 'CONTRIBUTION':
                header_key = f"CON_{rubric_ref.id}"
                order_val = rubric_ref.order if rubric_ref.order is not None else 999
                group_data['aportes_headers'][header_key] = {
                    'key': header_key, 'name': rubric_ref.name, 'order': order_val,
                    'abbreviation': rubric_ref.abbreviation
                }
                employee_dict['aportes_dict'][header_key] = employee_dict['aportes_dict'].get(header_key,
                                                                                              Decimal(0)) + val

            # B. Distribución de Partidas Presupuestarias
            budget_code = getattr(item, 'budget_line_code', None)
            if budget_code and str(budget_code).strip():
                if rubric_ref.has_mapping or (
                        item.item_type == 'INCOME' and 'REMUNERACION' in (rubric_ref.code or '').upper()):
                    budget_key = f"{budget_code}_{rubric_ref.name}"
                    if budget_key not in group_data['presupuesto']:
                        group_data['presupuesto'][budget_key] = {
                            'partida': budget_code, 'concepto': rubric_ref.name, 'monto': Decimal(0),
                            'order': rubric_ref.order or 99999
                        }
                    group_data['presupuesto'][budget_key]['monto'] += val

            # C. Distribución de la Matriz Contable Individual (Regla de Entrada de Cuentas)
            if spending_type.startswith('7'):
                debit_account = rubric_ref.debit_account_inv or rubric_ref.debit_account
                credit_account = rubric_ref.credit_account_inv or rubric_ref.credit_account
                group_data['es_inversion'] = True
            elif spending_type.startswith('6'):
                debit_account = rubric_ref.debit_account_prod or rubric_ref.debit_account
                credit_account = rubric_ref.credit_account_prod or rubric_ref.credit_account
            else:
                debit_account = rubric_ref.debit_account
                credit_account = rubric_ref.credit_account

            # Resolución dinámica de la cuenta puente contable (Pasivo)
            current_salary_rubric = _find_salary_rubric(spending_type)
            if current_salary_rubric:
                if spending_type.startswith('7'):
                    bridge_account = current_salary_rubric.credit_account_inv or current_salary_rubric.credit_account
                elif spending_type.startswith('6'):
                    bridge_account = current_salary_rubric.credit_account_prod or current_salary_rubric.credit_account
                else:
                    bridge_account = current_salary_rubric.credit_account
            else:
                bridge_account = None

            # 🛡️ ALERTAS DE CONFIGURACIÓN: Si falta una cuenta, creamos un indicador para mantener el balance
            if not debit_account:
                class DummyDebit:
                    code = f"FALTA_DEBE_{rubric_ref.code}"
                    name = f"[CONFIGURACIÓN] Falta cuenta DEBE para {rubric_ref.name}"
                    order = 99999

                debit_account = DummyDebit()

            if not credit_account:
                class DummyCredit:
                    code = f"FALTA_HABER_{rubric_ref.code}"
                    name = f"[CONFIGURACIÓN] Falta cuenta HABER para {rubric_ref.name}"
                    order = 99999

                credit_account = DummyCredit()

            if not bridge_account:
                class DummyBridge:
                    code = "FALTA_CUENTA_PUENTE"
                    name = f"[CONFIGURACIÓN] Falta cuenta de Pasivo/Puente para Sueldo ({spending_type})"
                    order = 99999

                bridge_account = DummyBridge()

            # Distribución por tipo de movimiento (Garantiza partida doble limpia)
            if item.item_type == 'INCOME':
                group_data['contabilidad'].setdefault(debit_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': debit_account.name,
                                                       'order': debit_account.order or 99999})
                group_data['contabilidad'][debit_account.code]['debe'] += val

                group_data['contabilidad'].setdefault(bridge_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': bridge_account.name,
                                                       'order': bridge_account.order or 99999})
                group_data['contabilidad'][bridge_account.code]['haber'] += val

            elif item.item_type == 'DEDUCTION':
                group_data['contabilidad'].setdefault(bridge_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': bridge_account.name,
                                                       'order': bridge_account.order or 99999})
                group_data['contabilidad'][bridge_account.code]['debe'] += val

                group_data['contabilidad'].setdefault(credit_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': credit_account.name,
                                                       'order': credit_account.order or 99999})
                group_data['contabilidad'][credit_account.code]['haber'] += val

                # Gestión del asiento de devengado simultáneo opcional (Asiento Espejo de Ingresos)
                if rubric_ref.income_account:
                    group_data['contabilidad'][credit_account.code]['debe'] += val
                    group_data['contabilidad'].setdefault(rubric_ref.income_account.code,
                                                          {'debe': Decimal(0), 'haber': Decimal(0),
                                                           'nombre': rubric_ref.income_account.name,
                                                           'order': rubric_ref.income_account.order or 99999})
                    group_data['contabilidad'][rubric_ref.income_account.code]['haber'] += val


            elif item.item_type == 'CONTRIBUTION':
                group_data['contabilidad'].setdefault(debit_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': debit_account.name,
                                                       'order': debit_account.order or 99999})
                group_data['contabilidad'][debit_account.code]['debe'] += val
                if bridge_account:
                    group_data['contabilidad'].setdefault(bridge_account.code,
                                                          {'debe': Decimal(0), 'haber': Decimal(0),
                                                           'nombre': bridge_account.name,
                                                           'order': bridge_account.order or 99999})
                    group_data['contabilidad'][bridge_account.code]['haber'] += val
                    group_data['contabilidad'][bridge_account.code]['debe'] += val
                group_data['contabilidad'].setdefault(credit_account.code,
                                                      {'debe': Decimal(0), 'haber': Decimal(0),
                                                       'nombre': credit_account.name,
                                                       'order': credit_account.order or 99999})
                group_data['contabilidad'][credit_account.code]['haber'] += val

        # ==============================================================
        # POST-PROCESO: LIQUIDACIÓN DE BANCOS Y CONTRAPARTIDAS PUENTE (UNIFICADO)
        # ==============================================================
        for group_key, group_data in report_data.items():
            for employee_id, employee_stuff in group_data['empleados'].items():

                for spending_code, net_value in employee_stuff['net_pay_by_spending_type'].items():
                    if net_value <= 0:
                        continue

                    salary_rubric_ref = _find_salary_rubric(spending_code)

                    bridge_account_bank = None
                    if salary_rubric_ref:
                        if spending_code.startswith('7'):
                            bridge_account_bank = salary_rubric_ref.credit_account_inv or salary_rubric_ref.credit_account
                        elif spending_code.startswith('6'):
                            bridge_account_bank = salary_rubric_ref.credit_account_prod or salary_rubric_ref.credit_account
                        else:
                            bridge_account_bank = salary_rubric_ref.credit_account

                    bank_account_obj = salary_rubric_ref.income_account if salary_rubric_ref else None

                    # 🛡️ Control de errores en asignación de Banco/Pasivo general
                    if not bridge_account_bank:
                        class DummyBridgeBank:
                            code = "FALTA_PUENTE_PAGO"
                            name = f"[CONFIGURACIÓN] Falta Pasivo/Puente de Sueldo para Pago ({spending_code})"
                            order = 99999

                        bridge_account_bank = DummyBridgeBank()

                    if not bank_account_obj:
                        class DummyBankAccount:
                            code = "FALTA_CUENTA_BANCO"
                            name = f"[CONFIGURACIÓN] Falta configurar cuenta de BANCO en Sueldo ({spending_code})"
                            order = 99999

                        bank_account_obj = DummyBankAccount()

                    group_data['contabilidad'].setdefault(bridge_account_bank.code,
                                                          {'debe': Decimal(0), 'haber': Decimal(0),
                                                           'nombre': bridge_account_bank.name,
                                                           'order': bridge_account_bank.order or 99999})
                    group_data['contabilidad'][bridge_account_bank.code]['debe'] += net_value

                    group_data['contabilidad'].setdefault(bank_account_obj.code,
                                                          {'debe': Decimal(0), 'haber': Decimal(0),
                                                           'nombre': bank_account_obj.name,
                                                           'order': bank_account_obj.order or 99999})
                    group_data['contabilidad'][bank_account_obj.code]['haber'] += net_value

            # Limpieza y ordenamiento final de las cabeceras de columnas del reporte
            group_data['ingresos_headers'] = {k: v for k, v in group_data['ingresos_headers'].items() if sum(
                emp['ingresos_dict'].get(k, Decimal(0)) for emp in group_data['empleados'].values()) > 0}
            group_data['descuentos_headers'] = {k: v for k, v in group_data['descuentos_headers'].items() if
                                                sum(emp['descuentos_dict'].get(k, Decimal(0)) for emp in
                                                    group_data['empleados'].values()) > 0}
            group_data['aportes_headers'] = {k: v for k, v in group_data['aportes_headers'].items() if sum(
                emp['aportes_dict'].get(k, Decimal(0)) for emp in group_data['empleados'].values()) > 0}

            group_data['ingresos_headers'] = sorted(group_data['ingresos_headers'].values(),
                                                    key=lambda x: (x['order'], x['name']))
            group_data['descuentos_headers'] = sorted(group_data['descuentos_headers'].values(),
                                                      key=lambda x: (x['order'], x['name']))
            group_data['aportes_headers'] = sorted(group_data['aportes_headers'].values(),
                                                   key=lambda x: (x['order'], x['name']))

            for emp in group_data['empleados'].values():
                emp['ingresos_list'] = [emp['ingresos_dict'].get(h['key'], Decimal(0)) for h in
                                        group_data['ingresos_headers']]
                emp['descuentos_list'] = [emp['descuentos_dict'].get(h['key'], Decimal(0)) for h in
                                          group_data['descuentos_headers']]
                emp['aportes_list'] = [emp['aportes_dict'].get(h['key'], Decimal(0)) for h in
                                       group_data['aportes_headers']]
                emp['total_aportes'] = sum(emp['aportes_dict'].values())

            totals_sabana = {
                'total_ingresos': sum((e['ingresos'] for e in group_data['empleados'].values()), Decimal(0)),
                'total_descuentos': sum((e['descuentos'] for e in group_data['empleados'].values()),
                                        Decimal(0)),
                'total_aportes': sum((e['total_aportes'] for e in group_data['empleados'].values()),
                                     Decimal(0)),
                'liquido': sum((e['liquido'] for e in group_data['empleados'].values()), Decimal(0)),
                'ingresos_list': [sum((e['ingresos_dict'].get(h['key'], Decimal(0)) for e in
                                       group_data['empleados'].values()), Decimal(0)) for h in
                                  group_data['ingresos_headers']],
                'descuentos_list': [sum((e['descuentos_dict'].get(h['key'], Decimal(0)) for e in
                                         group_data['empleados'].values()), Decimal(0)) for h in
                                    group_data['descuentos_headers']],
                'aportes_list': [
                    sum((e['aportes_dict'].get(h['key'], Decimal(0)) for e in group_data['empleados'].values()),
                        Decimal(0)) for h in group_data['aportes_headers']],
            }
            group_data['totales_sabana'] = totals_sabana

            group_data['colspans'] = {
                'ingresos': len(group_data['ingresos_headers']) or 1,
                'aportes': len(group_data['aportes_headers']),
                'descuentos': len(group_data['descuentos_headers'])
            }

            sorted_accounts_list = []
            for account_code_str, account_data in group_data['contabilidad'].items():
                if account_data['debe'] > 0 or account_data['haber'] > 0:
                    sorted_accounts_list.append({
                        'codigo': account_code_str, 'nombre': account_data['nombre'],
                        'debe': account_data['debe'], 'haber': account_data['haber'],
                        'order': account_data.get('order', 99999)
                    })

            group_data['contabilidad_ordenada'] = sorted(sorted_accounts_list,
                                                         key=lambda x: (x['order'], x['codigo']))
            group_data['total_contabilidad_debe'] = sum(
                (c['debe'] for c in group_data['contabilidad_ordenada']), Decimal(0))
            group_data['total_contabilidad_haber'] = sum(
                (c['haber'] for c in group_data['contabilidad_ordenada']), Decimal(0))

            # 🛠️ SALVACAÍDAS DE CUADRE: Corrige discrepancias infinitesimales de redondeo de centavos
            if group_data['total_contabilidad_debe'] != group_data['total_contabilidad_haber']:
                diff = group_data['total_contabilidad_debe'] - group_data['total_contabilidad_haber']
                adjust_account_code = "9.9.9.99"
                adjust_account_name = "[AJUSTE] Diferencia por Redondeo de Centavos"

                found = False
                for acc_row in sorted_accounts_list:
                    if acc_row['codigo'] == adjust_account_code:
                        if diff > 0:
                            acc_row['haber'] += diff
                        else:
                            acc_row['debe'] += abs(diff)
                        found = True
                        break

                if not found:
                    sorted_accounts_list.append({
                        'codigo': adjust_account_code, 'nombre': adjust_account_name,
                        'debe': Decimal(0) if diff > 0 else abs(diff),
                        'haber': diff if diff > 0 else Decimal(0),
                        'order': 99999
                    })

                group_data['contabilidad_ordenada'] = sorted(sorted_accounts_list,
                                                             key=lambda x: (x['order'], x['codigo']))
                group_data['total_contabilidad_debe'] = sum(
                    (c['debe'] for c in group_data['contabilidad_ordenada']), Decimal(0))
                group_data['total_contabilidad_haber'] = sum(
                    (c['haber'] for c in group_data['contabilidad_ordenada']), Decimal(0))

            group_data['total_presupuesto'] = sum((p['monto'] for p in group_data['presupuesto'].values()),
                                                  Decimal(0))
            group_data['presupuesto'] = dict(sorted(group_data['presupuesto'].items(),
                                                    key=lambda item: (int(item[1].get('order', 99999)),
                                                                      str(item[1].get('concepto', '')))))
            group_data['empleados'] = dict(sorted(group_data['empleados'].items(), key=lambda item: (
                    item[1]['persona'].last_name or "").lower()))

        sorted_reports = dict(sorted(report_data.items()))
        context = {
            'period': period,
            'grouped_reports': sorted_reports,
            'base_template': 'base_pdf.html' if request.GET.get('export') == 'pdf' else 'base.html'
        }
        return render(request, 'payroll/reports/grouped_financial_report.html', context)


class PayslipToggleWithholdView(LoginRequiredMixin, View):
    """
    API para encender/apagar la retención de pago de un rol específico.
    """

    def post(self, request, pk):
        payslip = get_object_or_404(Payslip, pk=pk)

        # Invertimos el estado actual
        payslip.is_withheld = not payslip.is_withheld
        payslip.save(update_fields=['is_withheld'])

        estado = "RETENIDO" if payslip.is_withheld else "LIBERADO"
        return JsonResponse({
            'success': True,
            'message': f'El pago de este empleado ha sido {estado}.',
            'is_withheld': payslip.is_withheld
        })


class PayslipItemUpdateAPIView(LoginRequiredMixin, View):
    """
    API para modificar un rubro manual (ej. subirle $10 a un descuento).
    Recalcula el total del empleado y RECONSTRUYE la contabilidad automáticamente.
    """

    def post(self, request, item_id):
        nuevo_valor = Decimal(request.POST.get('new_value', '0.00'))

        try:
            with transaction.atomic():
                item = get_object_or_404(PayslipItem, pk=item_id)
                item.value = nuevo_valor
                item.save(update_fields=['value'])

                # 1. RECALCULAR EL BOLSILLO DEL EMPLEADO
                payslip = item.payslip
                totales = PayslipItem.objects.filter(payslip=payslip).aggregate(
                    total_ing=Sum('value', filter=Q(item_type='INCOME')),
                    total_desc=Sum('value', filter=Q(item_type='DEDUCTION'))
                )

                t_ing = totales['total_ing'] or Decimal('0.00')
                t_desc = totales['total_desc'] or Decimal('0.00')

                payslip.total_income = t_ing
                payslip.total_deduction = t_desc
                payslip.net_pay = t_ing - t_desc
                payslip.save(update_fields=['total_income', 'total_deduction', 'net_pay'])

                # Ejecutar reconstrucción contable fuera de la transacción para evitar rollback
                def _safe_rebuild(pid):
                    try:
                        rebuild_accounting_for_period(pid)
                    except Exception as _e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.exception('Error al reconstruir contabilidad para periodo %s: %s', pid, _e)

                try:
                    transaction.on_commit(lambda: _safe_rebuild(payslip.period.id))
                except Exception:
                    # Fallback: intentar ejecutar sin on_commit
                    try:
                        _safe_rebuild(payslip.period.id)
                    except Exception:
                        pass

            return JsonResponse({
                'success': True,
                'message': 'Valor actualizado. Reconstrucción contable encolada.',
                'new_total_income': str(payslip.total_income),
                'new_total_deduction': str(payslip.total_deduction),
                'new_net_pay': str(payslip.net_pay)
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error en actualización: {str(e)}'})


class MarkPeriodAsPaidAPIView(LoginRequiredMixin, View):
    """
    Sella los roles actuales como PAGADOS para que no vuelvan a salir en futuros alcances.
    Si detecta que ya no queda nadie por cobrar, cierra el periodo automáticamente.
    """

    def post(self, request, period_id):
        # 1. Marcamos como pagados SOLO a los que no están retenidos y que aún no cobraban
        roles_actualizados = Payslip.objects.filter(
            period_id=period_id,
            is_withheld=False,
            is_paid=False
        ).update(is_paid=True)

        # 2. Verificamos si en este periodo todavía queda alguien sin cobrar
        faltan_por_pagar = Payslip.objects.filter(period_id=period_id, is_paid=False).exists()

        mensaje = f'Se han sellado {roles_actualizados} roles como PAGADOS en el SPI-SP.'

        if not faltan_por_pagar:
            # 3. ¡Magia! Si ya no hay nadie pendiente, cerramos el periodo para siempre
            PayrollPeriod.objects.filter(id=period_id).update(is_closed=True)
            mensaje += ' Como ya no quedan pagos pendientes, el Periodo se ha CERRADO automáticamente.'

        return JsonResponse({
            'success': True,
            'message': mensaje,
            'is_closed': not faltan_por_pagar
        })


class RecalculatePayslipsView(LoginRequiredMixin, View):
    """
    Recalcula los roles que coinciden con los filtros enviados (q, group, show_withheld).
    Se usa desde la interfaz para recalcular la búsqueda actual sin regenerar todo el periodo.
    """

    def post(self, request):
        period_id = request.POST.get('period_id') or request.GET.get('period_id')
        if not period_id:
            return JsonResponse({'success': False, 'message': 'Periodo no especificado.'}, status=400)

        period = get_object_or_404(PayrollPeriod, pk=period_id)

        # Capturar filtros opcionales
        search_query = (request.POST.get('q') or request.GET.get('q') or '').strip()
        group_filter = (request.POST.get('group') or request.GET.get('group') or '').strip()
        regime_filter = (request.POST.get('regime') or request.GET.get('regime') or '').strip()  # NUEVO
        show_withheld = (request.POST.get('show_withheld') or request.GET.get('show_withheld') or '').lower()

        qs = Payslip.objects.filter(period=period)
        if search_query:
            qs = qs.filter(
                Q(employee__person__first_name__icontains=search_query) |
                Q(employee__person__last_name__icontains=search_query) |
                Q(employee__person__document_number__icontains=search_query) |
                Q(items__budget_line__budget_group__short_code__icontains=search_query)
            ).distinct()
        if group_filter:
            qs = qs.filter(items__budget_line__budget_group__short_code=group_filter).distinct()

        if show_withheld in ['only', '1', 'true', 'yes']:
            qs = qs.filter(is_withheld=True)
        if regime_filter:
            qs = qs.filter(items__budget_line__regime_item_id=regime_filter).distinct()

        emp_ids = list(qs.values_list('employee_id', flat=True).distinct())
        if not emp_ids:
            return JsonResponse({'success': True, 'message': 'No hay roles que coincidan con los filtros.', 'count': 0})

        employees = list(Employee.objects.filter(id__in=emp_ids))

        seen_emp_ids = set()
        pairs = []
        for p in Payslip.objects.filter(period=period, employee_id__in=emp_ids).order_by('employee_id', 'id'):
            if p.employee_id not in seen_emp_ids:
                seen_emp_ids.add(p.employee_id)
                pairs.append((p.employee, p.worked_days or period.working_days))
        try:
            svc = PayrollCalculatorService(period, employees)
            result = svc.generate_for_selected(pairs)
            return JsonResponse(
                {'success': True, 'message': 'Recalculo ejecutado.', 'result': result, 'count': len(emp_ids)})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)


class GenerateMissingPayrollView(LoginRequiredMixin, View):
    """
    Busca empleados que tengan un contrato activo en el mes pero que AÚN NO
    tengan un rol de pagos generado, y los agrega sin borrar al resto.
    """

    def post(self, request):
        try:
            period_id = request.POST.get('period_id')
            period = get_object_or_404(PayrollPeriod, pk=period_id)

            # 1. ¿Quiénes DEBERÍAN estar en este mes? (Máquina del tiempo)
            valid_history_emp_ids = set(BudgetAssignmentHistory.objects.filter(
                start_date__lte=period.end_date
            ).filter(
                Q(end_date__isnull=True) | Q(end_date__gte=period.start_date)
            ).values_list('employee_id', flat=True))

            # 2. ¿Quiénes YA ESTÁN en el rol actual?
            existing_emp_ids = set(Payslip.objects.filter(period=period).values_list('employee_id', flat=True))

            # 3. Matemática de conjuntos: Los que deberían estar MENOS los que ya están = Los Nuevos
            missing_ids = valid_history_emp_ids - existing_emp_ids

            if not missing_ids:
                return JsonResponse(
                    {'status': 'info', 'message': 'Todos los empleados activos ya están en el rol. No hay faltantes.'})

            # 4. Traemos a los empleados y los mandamos al motor
            missing_employees = Employee.objects.filter(
                id__in=missing_ids,
            ).select_related('person', 'person__economic_data', 'person__economic_data__payroll_info')
            employees_with_days = [(emp, period.working_days) for emp in missing_employees]

            service = PayrollCalculatorService(period, missing_employees)
            # Usamos la función que calcula solo a los seleccionados
            resultado = service.generate_for_selected(employees_with_days)

            if resultado.get("success"):
                return JsonResponse({'status': 'success',
                                     'message': f'Se agregaron {len(missing_employees)} nuevos empleados al rol exitosamente.'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Hubo advertencias al calcular.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class BankTransferReportView(LoginRequiredMixin, View):
    """
    Vista EXCLUSIVA para el Archivo de Transferencias Bancarias (Reporte 4 / SPI-SP).
    """

    def get(self, request, pk):
        period = get_object_or_404(PayrollPeriod, pk=pk)
        tipo_filtro = request.GET.get('filtro', 'NORMAL')
        search_query = request.GET.get('q', '').strip()
        group_filter = request.GET.get('group', '').strip()
        regime_filter = request.GET.get('regime', '').strip()

        payslips_qs = Payslip.objects.filter(period=period).select_related(
            'employee__person'
        ).order_by('employee__person__last_name')

        # 1. APLICAMOS EL MISMO BUSCADOR DE LA PANTALLA PRINCIPAL
        if search_query:
            payslips_qs = payslips_qs.filter(
                Q(employee__person__first_name__icontains=search_query) |
                Q(employee__person__last_name__icontains=search_query) |
                Q(employee__person__document_number__icontains=search_query) |
                Q(items__budget_line__budget_group__short_code__icontains=search_query)
            ).distinct()  # Importante el distinct

        if group_filter:
            payslips_qs = payslips_qs.filter(items__budget_line__budget_group__short_code=group_filter).distinct()

        if regime_filter:
            payslips_qs = payslips_qs.filter(items__budget_line__regime_item_id=regime_filter).distinct()

        # 2. Filtros de Retención (Liberados vs Retenidos)
        # Soporte para parámetro `show_withheld` enviado por el frontend
        show_withheld = (self.request.GET.get('show_withheld') or '').lower()
        if show_withheld in ['only', '1', 'true', 'yes']:
            payslips_qs = payslips_qs.filter(is_withheld=True)
        else:
            if tipo_filtro == 'NORMAL':
                payslips_qs = payslips_qs.filter(is_withheld=False)
            elif tipo_filtro == 'REZAGADOS':
                payslips_qs = payslips_qs.filter(is_withheld=False, is_paid=False)

        if regime_filter:
            payslips_qs = payslips_qs.filter(items__budget_line__regime_item_id=regime_filter).distinct()

        # 3. PRE-CARGA CRÍTICA: Traemos los datos bancarios para evitar que el template colapse
        payslips_qs = payslips_qs.prefetch_related(
            'employee__person__economic_data__bank_account__bank',
            'employee__person__economic_data__bank_account__account_type'
        )

        total_transferir = sum(Decimal(str(p.net_pay)) for p in payslips_qs)

        context = {
            'period': period,
            'payslips': payslips_qs,
            'total_transferir': total_transferir,
            'filtro': tipo_filtro,
            'search_query': search_query
        }
        return render(request, 'payroll/reports/bank_transfer_report.html', context)


class PeriodUpdateView(UpdateView):
    model = PayrollPeriod
    form_class = PayrollPeriodForm
    template_name = 'payroll/modals/modal_period_form.html'

    def form_valid(self, form):
        self.object = form.save()
        return JsonResponse({'status': 'success', 'message': 'Periodo actualizado correctamente.'})

    def form_invalid(self, form):
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


def api_calculate_working_days(request):
    month_name = request.GET.get('month')
    year = request.GET.get('year')

    if not month_name or not year:
        return JsonResponse({'status': 'error', 'message': 'Faltan parámetros'}, status=400)

    try:
        # Mapeo de meses
        months_map = {
            'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
            'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
            'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
        }
        month_num = months_map.get(month_name.upper())
        year_num = int(year)

        # Calcular primer y último día del mes
        first_day = date(year_num, month_num, 1)
        last_day_num = calendar.monthrange(year_num, month_num)[1]
        last_day = date(year_num, month_num, last_day_num)

        # Usamos un objeto temporal del modelo para aprovechar la lógica de feriados
        temp_period = PayrollPeriod(start_date=first_day, end_date=last_day)
        working_days = temp_period.get_working_days()  # Esta función ya existe en tu models.py

        # Además, detectamos si hay feriados activos en el rango y construimos una advertencia
        from schedule.models import ScheduleObservation

        holidays = ScheduleObservation.objects.filter(
            is_holiday=True,
            is_active=True,
            start_date__lte=last_day,
            end_date__gte=first_day
        )

        # Compilar conjunto de fechas de feriados
        holiday_dates = set()
        from datetime import timedelta
        for holiday in holidays:
            curr = max(holiday.start_date, first_day)
            end_limit = min(holiday.end_date, last_day)
            while curr <= end_limit:
                holiday_dates.add(curr)
                curr += timedelta(days=1)

        response = {
            'status': 'success',
            'start_date': first_day.strftime('%Y-%m-%d'),
            'end_date': last_day.strftime('%Y-%m-%d'),
            'working_days': working_days
        }

        if len(holiday_dates) > 0:
            # Formatear una advertencia legible
            sample_dates = ', '.join(sorted([d.strftime('%d/%m/%Y') for d in list(holiday_dates)[:3]]))
            more = '' if len(holiday_dates) <= 3 else f' y {len(holiday_dates) - 3} más'
            response['warning'] = f'Se detectaron {len(holiday_dates)} feriado(s) en el periodo ({sample_dates}{more}).'
        else:
            response['info'] = 'No se detectaron feriados en el periodo seleccionado.'

        return JsonResponse(response)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def export_negative_balances_report(request, period_id):
    """
    Genera la vista de los empleados cuyo sueldo no alcanzó
    para cubrir sus descuentos, agrupado por tipo de descuento.
    """
    period = get_object_or_404(PayrollPeriod, id=period_id)

    debts = PendingDebt.objects.filter(
        period=period,
        pending_balance__gt=0
    ).select_related(
        'employee__person', 'rubric'
    ).order_by('rubric__name', 'employee__person__last_name')

    grouped_data = {}
    total_original = Decimal('0.0')
    total_cobrado = Decimal('0.0')
    total_pendiente = Decimal('0.0')

    for debt in debts:
        concept_name = debt.rubric.name
        if concept_name not in grouped_data:
            grouped_data[concept_name] = {
                'items': [],
                'sub_original': Decimal('0.0'),
                'sub_cobrado': Decimal('0.0'),
                'sub_pendiente': Decimal('0.0'),
            }

        grouped_data[concept_name]['items'].append(debt)
        grouped_data[concept_name]['sub_original'] += debt.original_value
        grouped_data[concept_name]['sub_cobrado'] += debt.collected_value
        grouped_data[concept_name]['sub_pendiente'] += debt.pending_balance

        total_original += debt.original_value
        total_cobrado += debt.collected_value
        total_pendiente += debt.pending_balance

    context = {
        'period': period,
        'grouped_data': grouped_data,
        'total_original': total_original,
        'total_cobrado': total_cobrado,
        'total_pendiente': total_pendiente,
        'has_debts': debts.exists(),
        # Variable opcional por si decides implementar un botón de imprimir en el HTML
        'auto_print': False
    }

    # Renderizamos directamente el HTML, igual que en el Reporte Bancario
    return render(request, 'payroll/reports/report_negative_balances.html', context)


class MassUpdateReserveFundsView(View):
    """
    API para carga masiva de Fondos de Reserva.
    Lee un Excel con Cédulas:
    - Los que estén en el Excel pasan a ACUMULAR (False).
    - Los que NO estén pasan a MENSUALIZAR (True) automáticamente.
    """

    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({'status': 'error', 'message': 'No se subió ningún archivo'})

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            # 1. Leer las cédulas del Excel (ignora si la primera celda dice "CEDULA")
            cedulas_acumulan = set()
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not row[0]: continue
                raw = str(row[0]).strip()
                if raw.replace('.', '').isdigit():
                    if raw.endswith('.0'): raw = raw[:-2]
                    cedulas_acumulan.add(raw.zfill(10))

            # 2. Traer a los empleados y preparar la actualización masiva
            empleados = Employee.objects.filter(is_active=True).select_related('person__economic_data__payroll_info')
            infos_to_update = []

            with transaction.atomic():
                for emp in empleados:
                    try:
                        # Navegamos hasta el perfil económico/nómina del empleado
                        info = emp.person.economic_data.payroll_info
                        if info:
                            cedula = emp.person.document_number

                            # LÓGICA MAESTRA: Si está en el Excel, Acumula (False). Si no, Mensualiza (True).
                            nuevo_estado = False if cedula in cedulas_acumulan else True

                            # Solo lo actualizamos si es diferente a lo que ya tenía, para ahorrar memoria
                            if getattr(info, 'reserve_funds') != nuevo_estado:
                                info.reserve_funds = nuevo_estado
                                infos_to_update.append(info)
                    except AttributeError:
                        continue

                # 3. Guardado ultra-rápido en base de datos (Bulk Update)
                if infos_to_update:
                    ModelClass = type(infos_to_update[0])  # Obtenemos el modelo dinámicamente
                    ModelClass.objects.bulk_update(infos_to_update, ['reserve_funds'])

            return JsonResponse({
                'status': 'success',
                'message': f'¡Proceso exitoso! Se configuraron {len(cedulas_acumulan)} empleados para ACUMULAR en el IESS. El resto mensualizará automáticamente.'
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error procesando el archivo: {str(e)}'})


class PrintablePayslipView(DetailView):
    """Genera el rol de pagos en pantalla completa para impresión oficial - ACCESO PÚBLICO CON VALIDACIÓN POR TOKEN"""
    model = Payslip
    template_name = 'payroll/reports/printable_payslip.html'
    context_object_name = 'payslip'

    def get_object(self, queryset=None):
        """Permite acceso público si viene validado por token en parámetro GET"""
        pk = self.kwargs.get('pk')
        token = self.request.GET.get('token')
        payslip = get_object_or_404(Payslip, pk=pk)

        # Si hay token, validamos que sea válido para este payslip
        if token:
            try:
                payslip_id = parse_public_payslip_token(token)
                if payslip_id != payslip.id:
                    raise Http404('Código de validación no coincide con el rol')
            except (signing.BadSignature, ValueError, TypeError):
                raise Http404('Código de validación inválido o expirado')

        return payslip

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payslip = self.object

        if num2words:
            try:
                entero = int(payslip.net_pay)
                decimales = int(round((payslip.net_pay - entero) * 100))
                letras_entero = num2words(entero, lang='es').upper()
                context['net_pay_words'] = f"{letras_entero} CON {decimales:02d}/100 DÓLARES"
            except Exception:
                context['net_pay_words'] = f"{payslip.net_pay} (Error al convertir)"
        else:
            context['net_pay_words'] = f"{payslip.net_pay} (Instalar num2words)"

        # Ordenamos igual que en el modal
        context['incomes'] = self.object.items.filter(item_type='INCOME').order_by('rubric__order')
        context['deductions'] = self.object.items.filter(item_type='DEDUCTION').order_by('rubric__order')

        validation_token = build_public_payslip_token(payslip.id)
        validation_url = self.request.build_absolute_uri(
            reverse('payroll:payslip_public_validate', kwargs={'token': validation_token})
        )

        import qrcode as qr_module
        qr = qr_module.QRCode(
            version=1,
            error_correction=qr_module.constants.ERROR_CORRECT_L,
            box_size=10,
            border=2,
        )
        qr.add_data(validation_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')

        buffered = io.BytesIO()
        img.save(buffered, format='PNG')
        context['qr_code'] = base64.b64encode(buffered.getvalue()).decode()
        context['validation_code'] = f"{validation_token[:10]}...{validation_token[-8:]}"
        context['validation_url'] = validation_url
        context['auto_print'] = True
        return context


class PublicPayslipValidationView(DetailView):
    """Vista publica para validar un rol de pago mediante token QR."""
    model = Payslip
    template_name = 'payroll/reports/printable_payslip.html'
    context_object_name = 'payslip'

    def get_object(self, queryset=None):
        token = self.kwargs.get('token')
        try:
            payslip_id = parse_public_payslip_token(token)
        except (signing.BadSignature, ValueError, TypeError):
            raise Http404('Codigo de validacion invalido')
        return get_object_or_404(Payslip, pk=payslip_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payslip = self.object
        token = self.kwargs.get('token')

        if num2words:
            try:
                entero = int(payslip.net_pay)
                decimales = int(round((payslip.net_pay - entero) * 100))
                letras_entero = num2words(entero, lang='es').upper()
                context['net_pay_words'] = f"{letras_entero} CON {decimales:02d}/100 DOLARES"
            except Exception:
                context['net_pay_words'] = f"{payslip.net_pay} (Error al convertir)"
        else:
            context['net_pay_words'] = f"{payslip.net_pay} (Instalar num2words)"

        context['incomes'] = self.object.items.filter(item_type='INCOME').order_by('rubric__order')
        context['deductions'] = self.object.items.filter(item_type='DEDUCTION').order_by('rubric__order')
        validation_url = self.request.build_absolute_uri(
            reverse('payroll:payslip_public_validate', kwargs={'token': token})
        )

        import qrcode as qr_module
        qr = qr_module.QRCode(
            version=1,
            error_correction=qr_module.constants.ERROR_CORRECT_L,
            box_size=10,
            border=2,
        )
        qr.add_data(validation_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')

        buffered = io.BytesIO()
        img.save(buffered, format='PNG')
        context['qr_code'] = base64.b64encode(buffered.getvalue()).decode()
        context['validation_code'] = f"{token[:10]}...{token[-8:]}"
        context['validation_url'] = validation_url
        context['auto_print'] = False
        return context


class SendPayslipEmailView(LoginRequiredMixin, View):
    """Genera PDF usando el mismo template que la impresión y lo envía por correo"""

    def post(self, request, pk):
        payslip = get_object_or_404(Payslip, pk=pk)

        inst_data = getattr(payslip.employee, 'institutional_data', None)
        correo_empleado = inst_data.institutional_email if inst_data else None

        if not correo_empleado:
            return JsonResponse({'status': 'error', 'message': 'Sin correo institucional.'})

        try:
            mes_str = str(payslip.period.month).upper()
            asunto = f"Notificación de Pago: {mes_str} {payslip.period.year} - GAD Loja"

            context_email = {
                'empleado': payslip.employee.person.full_name,
                'mes': payslip.period.month,
                'anio': payslip.period.year,
                'liquido': payslip.net_pay,
            }
            html_content = render_to_string('payroll/emails/payslip_notification.html', context_email)
            text_content = strip_tags(html_content)

            remitente = getattr(settings, 'DEFAULT_FROM_EMAIL', 'nomina@loja.gob.ec')
            email = EmailMultiAlternatives(subject=asunto, body=text_content, from_email=remitente,
                                           to=[correo_empleado])
            email.attach_alternative(html_content, "text/html")

            # ====================================================================
            # GENERAR PDF DESDE EL MISMO TEMPLATE QUE LA IMPRESIÓN
            # ====================================================================
            # Preparar el contexto igual que PrintablePayslipView
            context = {
                'payslip': payslip,
                'incomes': payslip.items.filter(item_type='INCOME').order_by('rubric__order'),
                'deductions': payslip.items.filter(item_type='DEDUCTION').order_by('rubric__order'),
            }

            # Generar números en letras
            if num2words:
                try:
                    entero = int(payslip.net_pay)
                    decimales = int(round((payslip.net_pay - entero) * 100))
                    letras_entero = num2words(entero, lang='es').upper()
                    context['net_pay_words'] = f"{letras_entero} CON {decimales:02d}/100 DÓLARES"
                except Exception:
                    context['net_pay_words'] = f"{payslip.net_pay} (Error al convertir)"
            else:
                context['net_pay_words'] = f"{payslip.net_pay} (Instalar num2words)"

            # Generar QR de validación
            validation_token = build_public_payslip_token(payslip.id)
            validation_url = request.build_absolute_uri(
                reverse('payroll:payslip_public_validate', kwargs={'token': validation_token})
            )

            import qrcode as qr_module
            qr = qr_module.QRCode(
                version=1,
                error_correction=qr_module.constants.ERROR_CORRECT_L,
                box_size=10,
                border=2,
            )
            qr.add_data(validation_url)
            qr.make(fit=True)
            img = qr.make_image(fill_color='black', back_color='white')

            buffered = io.BytesIO()
            img.save(buffered, format='PNG')
            context['qr_code'] = base64.b64encode(buffered.getvalue()).decode()
            context['validation_code'] = f"{validation_token[:10]}...{validation_token[-8:]}"
            context['validation_url'] = validation_url
            context['auto_print'] = False

            # Renderizar el template HTML a string
            html_payslip = render_to_string('payroll/reports/printable_payslip.html', context)

            # Convertir HTML a PDF usando xhtml2pdf
            pdf_buffer = io.BytesIO()
            pisa.CreatePDF(
                io.StringIO(html_payslip),
                pdf_buffer,
                raise_exception=True
            )
            pdf_file = pdf_buffer.getvalue()
            pdf_buffer.close()

            # Adjuntar PDF al correo
            nombre_archivo = f"Rol_{mes_str}_{payslip.period.year}.pdf"
            email.attach(nombre_archivo, pdf_file, 'application/pdf')
            email.send(fail_silently=False)

            return JsonResponse(
                {'status': 'success', 'message': f'Notificación enviada exitosamente a {correo_empleado}'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error al enviar por Zimbra: {str(e)}'})
