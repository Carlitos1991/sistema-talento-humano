from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.cache import cache
from django.db.models import Q, F
from django.db.models.functions import Length
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView, View, DetailView
from employee.models import Employee
from .forms import AdministrativeUnitForm, OrganizationalLevelForm, DeliverableForm, OrganigramForm
from .forms import AssignBossForm
from .models import AdministrativeUnit, OrganizationalLevel, Deliverable, InstitutionOrganigram


def _get_reassigned_unit_names_for_boss(boss, exclude_unit_id=None):
    """Obtiene los nombres de unidades donde el jefe ya estaba asignado."""
    if not boss:
        return []

    qs = AdministrativeUnit.objects.filter(boss=boss)
    if exclude_unit_id:
        qs = qs.exclude(pk=exclude_unit_id)

    return list(qs.order_by('name').values_list('name', flat=True))


class ParentOptionsJsonView(LoginRequiredMixin, View):
    def get(self, request):
        level_id = request.GET.get('level_id')
        direct_parent_only = request.GET.get('direct_parent_only', 'false').lower() == 'true'

        if not level_id or not str(level_id).isdigit():
            return JsonResponse({'results': []})

        try:
            current_level = OrganizationalLevel.objects.get(pk=int(level_id))

            if current_level.level_order <= 1:
                return JsonResponse({'results': []})

            # Si direct_parent_only=true, solo traer del nivel inmediatamente anterior
            if direct_parent_only:
                parent_level_order = current_level.level_order - 1
                parents = AdministrativeUnit.objects.filter(
                    level__level_order=parent_level_order,
                    is_active=True
                ).select_related('level').order_by('name')
            else:
                # Traer de todos los niveles anteriores (para casos de edición)
                parents = AdministrativeUnit.objects.filter(
                    level__level_order__lt=current_level.level_order,
                    is_active=True
                ).select_related('level').order_by('level__level_order', 'name')

            results = [{'id': p.id, 'text': f"{p.name} ➝ {p.level.name}"} for p in parents]
            return JsonResponse({'results': results})

        except OrganizationalLevel.DoesNotExist:
            return JsonResponse({'results': []})


class EmployeeSearchJsonView(LoginRequiredMixin, View):
    def get(self, request):
        term = request.GET.get('term', '').strip()
        qs = Employee.objects.filter(is_active=True).select_related('person')

        if term:
            qs = qs.filter(
                Q(person__first_name__icontains=term) |
                Q(person__last_name__icontains=term) |
                Q(person__document_number__icontains=term)
            )

        qs = qs[:20]
        results = []
        for emp in qs:
            full_name = f"{emp.person.last_name} {emp.person.first_name}"
            document = emp.person.document_number
            results.append({'id': str(emp.id), 'text': f"{full_name} ({document})"})

        return JsonResponse({'results': results})


# --- ESTADÍSTICAS ---
def get_unit_stats():
    return {
        'total': AdministrativeUnit.objects.count(),
        'active': AdministrativeUnit.objects.filter(is_active=True).count(),
        'inactive': AdministrativeUnit.objects.filter(is_active=False).count(),
    }


# --- LISTA ---
class UnitListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = AdministrativeUnit
    template_name = 'institution/unit_list.html'
    context_object_name = 'units'
    permission_required = 'institution.view_administrativeunit'

    def get_queryset(self):
        qs = AdministrativeUnit.objects.all().select_related(
            'level', 'parent', 'boss__person'
        ).annotate(
            code_len=Length('code')
        ).order_by('level__level_order', 'code_len', 'code', 'name')

        q = self.request.GET.get('q')
        show_inactive = self.request.GET.get('show_inactive')

        # Filtrar por nivel raíz (nivel 1) y por activo/inactivo según parámetro
        qs = qs.filter(level__level_order=1)
        if show_inactive == 'true':
            qs = qs.filter(is_active=False)
        else:
            qs = qs.filter(is_active=True)

        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = AdministrativeUnitForm()
        # Removed level stats and global counts to improve performance
        return context

    def get(self, request, *args, **kwargs):
        # Soporte AJAX para recargar solo la tabla parcial
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            html = render_to_string(
                'institution/partials/partial_unit_table.html',
                context,
                request=request
            )
            return JsonResponse({'html': html})
        return super().get(request, *args, **kwargs)


# --- CREAR ---
class UnitCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = AdministrativeUnit
    form_class = AdministrativeUnitForm
    template_name = 'institution/modals/modal_unit_form.html'
    permission_required = 'institution.add_administrativeunit'

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            selected_boss = form.cleaned_data.get('boss')
            reassigned_from = _get_reassigned_unit_names_for_boss(selected_boss)

            unit = form.save(commit=False)
            unit.is_active = True
            unit.save()

            message = 'Unidad creada correctamente.'
            if reassigned_from:
                message += f" El jefe fue reasignado automáticamente desde: {', '.join(reassigned_from)}."

            return JsonResponse({
                'success': True,
                'message': message,
                'new_stats': get_unit_stats()
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


# --- DETALLES (HTML) ---
class UnitDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = AdministrativeUnit
    template_name = 'institution/institution_unit_detail.html'
    context_object_name = 'unit'
    permission_required = 'institution.view_administrativeunit'

    def get_all_descendant_units(self, unit):
        """Obtiene recursivamente todos los IDs de descendientes de una unidad"""
        descendants = [unit.pk]
        children = AdministrativeUnit.objects.filter(parent=unit, is_active=True)
        for child in children:
            descendants.extend(self.get_all_descendant_units(child))
        return descendants

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_unit = self.object
        children = AdministrativeUnit.objects.filter(
            parent=current_unit, is_active=True
        ).order_by('code', 'name')

        # Empleados solo de la unidad actual (para mostrar en la lista)
        employees = Employee.objects.filter(
            area_id=current_unit.pk, is_active=True
        ).exclude(
            Q(employment_status__name__icontains='EX EMPLEADO') |
            Q(employment_status__name__icontains='EX TRABAJADOR')
        ).select_related('person').prefetch_related(
            'current_budget_line', 'current_budget_line__position_item'
        )

        # Obtener IDs de todas las subdependencias (recursivo)
        all_unit_ids = self.get_all_descendant_units(current_unit)

        # Empleados de la unidad actual + todas sus subdependencias
        all_employees = Employee.objects.filter(
            area_id__in=all_unit_ids, is_active=True
        ).exclude(
            Q(employment_status__name__icontains='EX EMPLEADO') |
            Q(employment_status__name__icontains='EX TRABAJADOR')
        )
        # Propios: No tienen dependencia original registrada o es igual a su área actual
        propios_filter = Q(institutional_data__original_dependency=current_unit.pk) | \
                         Q(institutional_data__original_dependency__isnull=True, area_id=current_unit.pk)

        propios_count = all_employees.filter(propios_filter).count()

        reubicados_count = all_employees.filter(
            institutional_data__original_dependency__isnull=False
        ).exclude(
            institutional_data__original_dependency=F('area')
        ).count()

        # Estadísticas por estado laboral (incluyendo todas las subdependencias)
        from django.db.models import Count
        stats_qs = all_employees.values(
            'employment_status__code'
        ).annotate(total=Count('id'))

        stats_dict = {stat['employment_status__code']: stat['total']
                      for stat in stats_qs if stat['employment_status__code']}

        unit_stats = {
            'total': all_employees.count(),
            'empleado': stats_dict.get('EMPLEADO', 0),
            'trabajador': stats_dict.get('TRABAJADOR', 0),
            'contratado': stats_dict.get('CONTRATADO', 0),
            'propios': propios_count,
            'reubicados': reubicados_count,
        }

        context['children'] = children
        context['employees'] = employees  # Solo de la unidad actual para mostrar
        context['unit_stats'] = unit_stats  # Incluye todas las subdependencias
        context['form'] = AdministrativeUnitForm()
        return context


# --- DETALLES JSON (para modal de edición) ---


class UnitDetailJsonView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'institution.view_administrativeunit'

    def handle_no_permission(self):
        # Devuelve JSON en vez de redirigir a login o error HTML
        return JsonResponse({'success': False, 'error': 'Permiso denegado.'}, status=403)

    def get(self, request, pk):
        # Verificar permisos manualmente para AJAX
        if not request.user.has_perm(self.permission_required):
            return self.handle_no_permission()
        unit = get_object_or_404(AdministrativeUnit, pk=pk)
        boss_data = None
        if unit.boss:
            # Obtener cargo/posición del funcionario si existe
            position = ''
            try:
                cb = unit.boss.current_budget_line.first()
                if cb and getattr(cb, 'position_item', None):
                    position = cb.position_item.name
            except Exception:
                position = ''

            boss_data = {
                'id': unit.boss.id,
                'text': f"{unit.boss.person.first_name} {unit.boss.person.last_name}",
                'person_id': unit.boss.person.id,
                'photo_url': unit.boss.person.photo.url if getattr(unit.boss.person, 'photo', None) else '',
                'position': position,
            }
        data = {
            'name': unit.name,
            'level': unit.level_id,
            'parent': unit.parent_id,
            'parent_name': unit.parent.name if unit.parent else None,
            'parent_level': unit.parent.level_id if unit.parent else None,
            'boss': unit.boss_id,
            'boss_data': boss_data,
            'code': unit.code,
            'address': getattr(unit, 'address', '') or '',
            'phone': getattr(unit, 'phone', '') or '',
            'mission': getattr(unit, 'mission', '') or '',
            'is_active': unit.is_active,
        }
        return JsonResponse({'success': True, 'data': data})


# --- ACTUALIZAR ---
class UnitUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = AdministrativeUnit
    form_class = AdministrativeUnitForm
    template_name = 'institution/modals/modal_unit_form.html'
    permission_required = 'institution.change_administrativeunit'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        original_parent = self.object.parent
        original_level = self.object.level
        original_boss = self.object.boss

        form = self.get_form()
        if form.is_valid():
            selected_boss = form.cleaned_data.get('boss') or original_boss
            reassigned_from = _get_reassigned_unit_names_for_boss(selected_boss, exclude_unit_id=self.object.id)

            # Guardar explícitamente usando commit=False para asegurar que
            # los campos ocultos (parent/level) se apliquen correctamente.
            unit = form.save(commit=False)

            # Si el formulario no trajo algún campo, conservar el valor existente.
            if 'parent' in request.POST:
                unit.parent = form.cleaned_data.get('parent')
            else:
                unit.parent = original_parent

            if 'level' in request.POST and form.cleaned_data.get('level'):
                unit.level = form.cleaned_data.get('level')
            else:
                unit.level = original_level

            if 'boss' in request.POST:
                unit.boss = form.cleaned_data.get('boss')
            else:
                unit.boss = original_boss

            old_boss = original_boss
            unit.save()
            new_boss = unit.boss

            # Marcar nuevo jefe como tal
            if new_boss:
                new_boss.is_boss = True
                new_boss.save(update_fields=['is_boss'])

            # Si existía un jefe anterior distinto y ya no maneja unidades, desmarcarlo
            if old_boss and old_boss != new_boss and not old_boss.managed_units.exists():
                old_boss.is_boss = False
                old_boss.save(update_fields=['is_boss'])

            message = 'Unidad actualizada correctamente.'
            if reassigned_from:
                message += f" El jefe fue reasignado automáticamente desde: {', '.join(reassigned_from)}."

            return JsonResponse({'success': True, 'message': message})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


# --- CAMBIAR PADRE (REUBICAR) ---
class UnitChangeParentView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'institution.change_administrativeunit'

    def post(self, request, pk):
        unit = get_object_or_404(AdministrativeUnit, pk=pk)
        new_parent_id = request.POST.get('parent')

        try:
            # Validar que el nuevo padre sea válido (si no es vacío)
            if new_parent_id and str(new_parent_id).isdigit():
                new_parent = AdministrativeUnit.objects.get(pk=int(new_parent_id), is_active=True)

                # Validación: El nuevo padre no puede ser la misma unidad
                if int(new_parent_id) == unit.id:
                    return JsonResponse({
                        'success': False,
                        'message': 'No puedes asignar la unidad como su propio padre.',
                        'errors': {'parent': ['La unidad no puede ser su propio padre.']}
                    }, status=400)

                # Validación: El nuevo padre debe tener un nivel menor (order menor)
                if new_parent.level.level_order >= unit.level.level_order:
                    return JsonResponse({
                        'success': False,
                        'message': 'El padre debe tener un nivel jerárquico superior.',
                        'errors': {'parent': ['El padre debe tener un nivel jerárquico superior.']}
                    }, status=400)

                unit.parent = new_parent
            else:
                # Si no hay parent_id o es vacío, establecer como raíz
                unit.parent = None

            unit.save()

            return JsonResponse({
                'success': True,
                'message': f'Unidad "{unit.name}" reubicada correctamente.'
            })

        except AdministrativeUnit.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'La unidad padre seleccionada no existe o no está activa.',
                'errors': {'parent': ['Padre no válido.']}
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al reubicar la unidad: {str(e)}',
                'errors': {'__all__': [str(e)]}
            }, status=400)


# --- TOGGLE STATUS ---
@method_decorator(require_POST, name='dispatch')
class UnitToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'institution.change_administrativeunit'

    def post(self, request, pk):
        unit = get_object_or_404(AdministrativeUnit, pk=pk)
        unit.is_active = not unit.is_active
        cache.delete('level_stats')
        unit.save()

        status_label = "activada" if unit.is_active else "desactivada"
        return JsonResponse({
            'success': True,
            'message': f'La unidad "{unit.name}" ha sido {status_label}.',
        })


@login_required
def unit_partial_table(request):
    show_inactive = request.GET.get('show_inactive')
    parent_id = request.GET.get('parent_id')
    q = request.GET.get('q')

    # Base queryset with selects/annotations
    qs = AdministrativeUnit.objects.all().select_related('level', 'parent', 'boss__person') \
        .annotate(code_len=Length('code')) \
        .order_by('level__level_order', 'code_len', 'code', 'name')

    # Active/inactive filter
    if show_inactive == 'true':
        qs = qs.filter(is_active=False)
    else:
        qs = qs.filter(is_active=True)

    # Optional text search
    if q:
        qs = qs.filter(Q(name__icontains=q))
    elif parent_id:
        # Solo aplica navegación por dependencia cuando no hay búsqueda activa.
        qs = qs.filter(parent_id=parent_id)
    else:
        qs = qs.filter(level__level_order=1)

    context = get_units_context(units_queryset=qs)
    html = render_to_string(
        'institution/partials/partial_unit_table.html',
        context,
        request=request
    )
    return HttpResponse(html)


# ==========================================
# GESTIÓN DE NIVELES JERÁRQUICOS
# ==========================================

def get_level_stats():
    """
    Retorna niveles con conteo, colores e ICONOS dinámicos.
    """
    cached = cache.get('level_stats')
    if cached:
        return cached
    levels = OrganizationalLevel.objects.filter(is_active=True).order_by('level_order')[:5]
    stats = []
    total_count = AdministrativeUnit.objects.filter(is_active=True).count()
    stats.append({
        'id': 'total',
        'name': 'Total',
        'order': 0,
        'count': total_count,
        'color': 'color-zero',
        'icon': 'fa-layer-group'
    })

    colors = ['color-one', 'color-two', 'color-three', 'color-four', 'color-five']
    icons = ['fa-globe', 'fa-building', 'fa-briefcase', 'fa-users', 'fa-user-tag']

    for i, lvl in enumerate(levels):
        count = AdministrativeUnit.objects.filter(level=lvl, is_active=True).count()
        stats.append({
            'id': lvl.id,
            'name': lvl.name,
            'order': lvl.level_order,
            'count': count,
            'color': colors[i] if i < len(colors) else 'color-one',
            'icon': icons[i] if i < len(icons) else 'fa-sitemap'
        })

    result = {'level_stats': stats}
    cache.set('level_stats', result, timeout=300)  # 5 minutos
    return result


# ============================================================
# GESTIÓN DE NIVELES JERÁRQUICOS
# ============================================================
class LevelListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = OrganizationalLevel
    template_name = 'institution/level_list.html'
    context_object_name = 'levels'
    permission_required = 'institution.view_organizationallevel'

    def get_queryset(self):
        qs = OrganizationalLevel.objects.all().order_by('level_order')
        q = self.request.GET.get('q')
        show_inactive = self.request.GET.get('show_inactive')

        if q:
            qs = qs.filter(name__icontains=q)
            
        if show_inactive == 'true':
            qs = qs.filter(is_active=False)
        else:
            qs = qs.filter(is_active=True)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = OrganizationalLevelForm()
        context['total'] = OrganizationalLevel.objects.all().count()
        context['active'] = OrganizationalLevel.objects.filter(is_active=True).count()
        context['inactive'] = OrganizationalLevel.objects.filter(is_active=False).count()
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            html = render_to_string(
                'institution/partials/partial_level_table.html',
                context,
                request=request
            )
            stats = {
                'total': context.get('total', 0),
                'active': context.get('active', 0),
                'inactive': context.get('inactive', 0),
            }
            return JsonResponse({'html': html, 'stats': stats})
        return super().get(request, *args, **kwargs)


class LevelCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = OrganizationalLevel
    form_class = OrganizationalLevelForm
    template_name = 'institution/modals/modal_level_form.html'
    permission_required = 'institution.add_organizationallevel'

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            form.save()
            return JsonResponse({
                'success': True,
                'message': 'Nivel jerárquico creado.',
                'new_stats': get_level_stats()
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


class LevelDetailView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'institution.view_organizationallevel'

    def get(self, request, pk):
        lvl = get_object_or_404(OrganizationalLevel, pk=pk)
        return JsonResponse({
            'success': True,
            'data': {'name': lvl.name, 'level_order': lvl.level_order}
        })


class LevelUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = OrganizationalLevel
    form_class = OrganizationalLevelForm
    template_name = 'institution/modals/modal_level_form.html'
    permission_required = 'institution.change_organizationallevel'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Nivel actualizado correctamente.'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@require_POST
@permission_required('institution.change_organizationallevel', raise_exception=True)
def level_toggle_status(request, pk):
    lvl = get_object_or_404(OrganizationalLevel, pk=pk)
    next_status = not lvl.is_active

    if next_status:
        conflict = OrganizationalLevel.objects.filter(
            level_order=lvl.level_order,
            is_active=True
        ).exclude(pk=pk).exists()

        if conflict:
            return JsonResponse({
                'success': False,
                'message': f"Conflicto: Ya existe un nivel jerárquico #{lvl.level_order} activo. Desactive el anterior primero."
            }, status=400)

    lvl.is_active = next_status
    lvl.save()

    status_label = "activado" if lvl.is_active else "desactivado"
    return JsonResponse({
        'success': True,
        'message': f'Nivel "{lvl.name}" {status_label}.',
        'new_stats': get_level_stats()
    })


def api_get_administrative_children(request):
    parent_id = request.GET.get('parent_id')
    term = (request.GET.get('term') or '').strip()
    # Si se provee parent_id, traer hijos de ese padre.
    # Si no, limitar a raíces que sean de nivel 1 (nivel jerárquico superior).
    if parent_id:
        filters = {'parent_id': parent_id}
    else:
        # Sin parent_id: por compatibilidad devolvemos raíces.
        # Si hay búsqueda por término, permitimos buscar en todas las unidades activas.
        filters = {} if term else {'parent__isnull': True, 'level__level_order': 1}
    filters['is_active'] = True

    units = AdministrativeUnit.objects.filter(**filters).order_by('name')
    if term:
        units = units.filter(
            Q(name__icontains=term) |
            Q(code__icontains=term)
        )
    units = units[:30]

    data = []
    for u in units:
        has_children = u.children.filter(is_active=True).exists()
        data.append({'id': u.id, 'name': u.name, 'has_children': has_children})

    # Formato dual para compatibilidad: `units` (existente) + `results` (Select2)
    results = [{'id': str(item['id']), 'text': item['name']} for item in data]
    return JsonResponse({'success': True, 'units': data, 'results': results})


class DeliverableListJsonView(LoginRequiredMixin, View):
    def get(self, request, unit_id):
        deliverables = Deliverable.objects.filter(unit_id=unit_id, is_active=True).order_by('-created_at')
        data = [{'id': d.id, 'name': d.name, 'description': d.description or ''} for d in deliverables]
        return JsonResponse({'success': True, 'data': data})


class DeliverableCreateUpdateView(LoginRequiredMixin, View):
    def post(self, request, unit_id, pk=None):
        if pk:
            instance = get_object_or_404(Deliverable, pk=pk, unit_id=unit_id)
            form = DeliverableForm(request.POST, instance=instance)
        else:
            form = DeliverableForm(request.POST)

        if form.is_valid():
            deliverable = form.save(commit=False)
            deliverable.unit_id = unit_id
            deliverable.save()
            return JsonResponse({'success': True, 'message': 'Entregable guardado correctamente.'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


class DeliverableDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        deliverable = get_object_or_404(Deliverable, pk=pk)
        deliverable.is_active = False  # Soft delete
        deliverable.save()
        return JsonResponse({'success': True, 'message': 'Entregable eliminado.'})


@login_required
def api_unit_deliverables(request, unit_id):
    """
    Endpoint para obtener los entregables de una unidad administrativa.
    Utilizado por Vue.js en la vista de detalle y en el Wizard de Perfiles de Puesto.
    """
    try:
        deliverables = Deliverable.objects.filter(
            unit_id=unit_id,
            is_active=True
        ).order_by('name')
        data = [{'id': d.id, 'name': d.name, 'description': d.description or ''} for d in deliverables]
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


class GetNextCodeJsonView(LoginRequiredMixin, View):
    def get(self, request):
        parent_id = request.GET.get('parent_id')
        next_code = "1"
        suggested_level_id = None

        try:
            if parent_id and parent_id != 'null' and parent_id != 'undefined':
                # --- DEPENDENCIAS ---
                parent = get_object_or_404(AdministrativeUnit, pk=parent_id)
                parent_code = parent.code or "0"
                last_child = AdministrativeUnit.objects.filter(parent=parent).order_by('-id').first()

                if last_child and last_child.code:
                    try:
                        parts = last_child.code.split('.')
                        last_num = int(parts[-1])
                        next_code = f"{'.'.join(parts[:-1])}.{last_num + 1}"
                    except:
                        next_code = f"{parent_code}.1"
                else:
                    next_code = f"{parent_code}.1"

                lvl = OrganizationalLevel.objects.filter(level_order=(parent.level.level_order + 1)).first()
                if lvl: suggested_level_id = lvl.id

            else:
                # --- UNIDADES PADRE (NIVEL 1) ---
                # Buscamos el código numérico más alto entre las raíces
                from django.db.models.functions import Cast
                from django.db.models import IntegerField, Max

                max_code_root = AdministrativeUnit.objects.filter(
                    parent__isnull=True,
                    code__regex=r'^\d+$'  # Solo los que sean números
                ).annotate(
                    code_int=Cast('code', output_field=IntegerField())
                ).aggregate(max_val=Max('code_int'))['max_val']

                if max_code_root is not None:
                    next_code = str(max_code_root + 1)
                else:
                    next_code = "1"

                lvl = OrganizationalLevel.objects.filter(level_order=1).first()
                if lvl: suggested_level_id = lvl.id

            return JsonResponse({
                'success': True,
                'next_code': next_code,
                'suggested_level': suggested_level_id
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class OrganigramView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'institution.view_administrativeunit'
    template_name = 'institution/organigram.html'

    def get(self, request):
        organigram = InstitutionOrganigram.objects.last()
        form = OrganigramForm()
        return render(request, self.template_name, {'organigram': organigram, 'form': form})

    def post(self, request):
        if not request.user.has_perm('institution.change_administrativeunit'):
            return render(request, '403.html', status=403)

        form = OrganigramForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.updated_by = request.user
            instance.save()
            return redirect('institution:organigram_view')

        organigram = InstitutionOrganigram.objects.last()
        return render(request, self.template_name, {'organigram': organigram, 'form': form})


class RootLevelListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = AdministrativeUnit
    template_name = 'institution/root_level_list.html'
    context_object_name = 'units'
    permission_required = 'institution.view_administrativeunit'

    def get(self, request, pk=None):
        context = {}

        if pk:
            current_unit = get_object_or_404(AdministrativeUnit, pk=pk)
            children = AdministrativeUnit.objects.filter(
                parent=current_unit, is_active=True
            ).order_by('code', 'name')
            employees = Employee.objects.filter(
                area_id=current_unit.pk, is_active=True
            ).exclude(
                Q(employment_status__name__icontains='EX EMPLEADO') |
                Q(employment_status__name__icontains='EX TRABAJADOR')
            ).select_related('person').prefetch_related(
                'current_budget_line', 'current_budget_line__position_item'
            )
            context = {
                'mode': 'detail',
                'current_unit': current_unit,
                'children': children,
                'employees': employees,
            }
        else:
            units = AdministrativeUnit.objects.filter(
                level__level_order=1, is_active=True
            ).select_related('level').order_by('code', 'name')
            context = {
                'mode': 'root',
                'units': units,
                'total_active': units.count()
            }

        context['form'] = AdministrativeUnitForm()
        return render(request, self.template_name, context)


class UnitAssignBossView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = AdministrativeUnit
    form_class = AssignBossForm
    template_name = 'institution/modals/modal_assign_boss.html'
    permission_required = 'institution.change_administrativeunit'
    context_object_name = 'unit'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        old_boss = self.object.boss
        form = self.get_form()

        if form.is_valid():
            selected_boss = form.cleaned_data.get('boss')
            reassigned_from = _get_reassigned_unit_names_for_boss(selected_boss, exclude_unit_id=self.object.id)

            unit = form.save()
            new_boss = unit.boss

            if new_boss:
                new_boss.is_boss = True
                new_boss.save(update_fields=['is_boss'])

            if old_boss and old_boss != new_boss:
                if not old_boss.managed_units.exists():
                    old_boss.is_boss = False
                    old_boss.save(update_fields=['is_boss'])

            message = f'Jefe asignado correctamente a {self.object.name}.'
            if reassigned_from:
                message += f" Reasignado automáticamente desde: {', '.join(reassigned_from)}."

            return JsonResponse({'success': True, 'message': message})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
def level_partial_table(request):
    show_inactive = request.GET.get('show_inactive')
    if show_inactive == 'true':
        levels = OrganizationalLevel.objects.all().order_by('level_order')
    else:
        levels = OrganizationalLevel.objects.filter(is_active=True).order_by('level_order')
    context = {'levels': levels}
    html = render_to_string('institution/partials/partial_level_table.html', context, request=request)
    return HttpResponse(html)


def get_all_descendant_unit_ids(unit):
    """Función auxiliar para obtener recursivamente todos los IDs de descendientes"""
    descendants = [unit.pk]
    children = AdministrativeUnit.objects.filter(parent=unit, is_active=True)
    for child in children:
        descendants.extend(get_all_descendant_unit_ids(child))
    return descendants


@login_required
@permission_required('institution.view_administrativeunit', raise_exception=True)
@login_required
@permission_required('institution.view_administrativeunit', raise_exception=True)
def export_unit_employees_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    unit = get_object_or_404(AdministrativeUnit, pk=pk)
    status_code = request.GET.get('status', 'total')
    status_code_upper = status_code.upper() if status_code else 'TOTAL'

    all_unit_ids = get_all_descendant_unit_ids(unit)

    # Consulta base con select_related optimizado
    employees = Employee.objects.filter(
        area_id__in=all_unit_ids,
        is_active=True
    ).exclude(
        Q(employment_status__name__icontains='EX EMPLEADO') |
        Q(employment_status__name__icontains='EX TRABAJADOR')
    ).select_related(
        'person',
        'employment_status',
        'area',
        'institutional_data',
        'institutional_data__original_dependency'
    ).order_by('area__name', 'person__last_name')

    propios_filter = Q(institutional_data__original_dependency=unit.pk) | \
                     Q(institutional_data__original_dependency__isnull=True, area_id=unit.pk)

    if status_code_upper == 'PROPIOS':
        employees = employees.filter(propios_filter)
    elif status_code_upper == 'REUBICADOS':
        employees = employees.exclude(propios_filter)
    elif status_code_upper != 'TOTAL':
        employees = employees.filter(employment_status__code=status_code_upper)

    wb = Workbook()
    ws = wb.active
    ws.title = status_code_upper if status_code_upper != 'TOTAL' else "Todos"

    # Encabezados actualizados
    headers = ['N°', 'Apellidos y Nombres', 'Cédula/Documento', 'Dependencia Actual', 'Dependencia Original', 'Cargo',
               'Remuneración']
    ws.append(headers)

    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for idx, emp in enumerate(employees, start=1):
        full_name = f"{emp.person.last_name} {emp.person.first_name}"
        document = emp.person.document_number or '-'
        dependencia_actual = emp.area.name if emp.area else '-'

        # Determinar dependencia original
        if hasattr(emp, 'institutional_data') and emp.institutional_data.original_dependency:
            dependencia_orig = emp.institutional_data.original_dependency.name
        else:
            dependencia_orig = dependencia_actual

        cargo = '-'
        budget_line = emp.current_budget_line.first()
        if budget_line and budget_line.position_item:
            cargo = budget_line.position_item.name

        remuneracion = '-'
        if budget_line and hasattr(budget_line, 'remuneration') and budget_line.remuneration:
            remuneracion = f"${budget_line.remuneration:,.2f}"

        ws.append([idx, full_name, document, dependencia_actual, dependencia_orig, cargo, remuneracion])

    # Ajustar ancho de columnas (ahora son 7)
    column_widths = [8, 40, 18, 35, 35, 35, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Empleados_{unit.name.replace(' ', '_')}_{status_code_upper}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def get_units_context(units_queryset=None):
    """Auxiliar para mantener consistencia de contexto en vistas y AJAX"""
    if units_queryset is None:
        units_queryset = AdministrativeUnit.objects.all().select_related(
            'level', 'parent', 'boss__person'
        ).annotate(
            code_len=Length('code')
        ).order_by('level__level_order', 'code_len', 'code', 'name')
    return {
        'units': units_queryset,
    }
