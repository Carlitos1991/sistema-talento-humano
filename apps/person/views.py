# apps/person/views.py
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView, TemplateView
from django.db.models import Q, Value, CharField
from django.db.models.functions import Concat

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from budget.models import BudgetLine
from core.models import CatalogItem
from employee.models import Employee, InstitutionalData
from institution.models import AdministrativeUnit
from .models import Person, PersonAuditLog
from .forms import PersonForm
from .utils import log_person_audit, PERSON_AUDIT_SECTIONS


class EmployeeReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'person/employee_report.html'
    permission_required = 'person.view_person'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.urls import reverse
        context['units_api_url'] = reverse('institution:api_unit_children')
        context['export_url'] = reverse('person:employee_report_export')
        # Campos disponibles para el reporte (clave,label)
        # Agrupar campos en institucionales y personales para la UI
        context['available_fields'] = {
            'institutional': [
                ('area', 'Dependencia'),
                ('cargo', 'Cargo'),
                ('remuneration', 'Remuneración'),
            ],
            'personal': [
                ('blood_type', 'Tipo de Sangre'),
                ('marital_status', 'Estado Civil'),
                ('gender', 'Género'),
                ('birth_date', 'Fecha Nac.'),
                ('email', 'Correo Pers.'),
                ('address_reference', 'Dirección'),
                ('phone_number', 'Teléfono'),
                ('emergency_contact_name', 'Contacto Emerg.'),
                ('emergency_contact_phone', 'Celular Emerg.'),
            ]
        }
        return context


class EmployeeReportExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'person.view_person'

    def get_descendants_with_depth(self, unit, depth=0):
        result = [(unit, depth)]
        children = AdministrativeUnit.objects.filter(parent=unit, is_active=True).order_by('name')
        for child in children:
            result.extend(self.get_descendants_with_depth(child, depth + 1))
        return result

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        unit_id = request.GET.get('unit_id') or request.GET.get('pk')
        unit = None
        report_scope_name = 'TODA LA INSTITUCION'
        if unit_id:
            unit = get_object_or_404(AdministrativeUnit, pk=unit_id)
            report_scope_name = unit.name

        # fields como lista separada por comas
        fields_param = request.GET.get('fields', '')
        selected_fields = [f for f in (fields_param.split(',') if fields_param else []) if f]

        # Definir mapeo clave -> (label, getter lambda)
        def get_val(emp, key):
            if key == 'blood_type':
                return getattr(emp.person.blood_type, 'name', '') if emp.person and emp.person.blood_type else ''
            if key == 'marital_status':
                return getattr(emp.person.marital_status, 'name', '') if emp.person and emp.person.marital_status else ''
            if key == 'gender':
                return getattr(emp.person.gender, 'name', '') if emp.person and emp.person.gender else ''
            if key == 'birth_date':
                return emp.person.birth_date.strftime('%Y-%m-%d') if emp.person and emp.person.birth_date else ''
            if key == 'email':
                return emp.person.email or ''
            if key == 'address_reference':
                return emp.person.address_reference or ''
            if key == 'phone_number':
                return emp.person.phone_number or ''
            if key == 'area':
                return emp.area.name if emp.area else ''
            if key == 'cargo':
                bl = emp.current_budget_line.first()
                return bl.position_item.name if bl and getattr(bl, 'position_item', None) else ''
            if key == 'remuneration':
                bl = emp.current_budget_line.first()
                if bl and getattr(bl, 'remuneration', None):
                    try:
                        return f"${float(bl.remuneration):,.2f}"
                    except Exception:
                        return str(getattr(bl, 'remuneration', ''))
                return ''
            if key == 'emergency_contact_name':
                return emp.person.emergency_contact_name or ''
            if key == 'emergency_contact_phone':
                return emp.person.emergency_contact_phone or ''
            return ''

        # Obtener unidades con profundidad: si no se selecciona unidad,
        # incluir toda la institución desde las unidades raíz activas.
        if unit:
            units_with_depth = self.get_descendants_with_depth(unit, depth=0)
        else:
            units_with_depth = []
            root_units = AdministrativeUnit.objects.filter(parent__isnull=True, is_active=True).order_by('name')
            for root_unit in root_units:
                units_with_depth.extend(self.get_descendants_with_depth(root_unit, depth=0))

        # Preparar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Empleados'

        # Columnas: siempre N°, Apellidos, Nombres, Documento
        institutional_keys = ['area', 'cargo', 'remuneration']
        personal_keys = ['blood_type', 'marital_status', 'gender', 'birth_date', 'email', 'address_reference', 'phone_number', 'emergency_contact_name', 'emergency_contact_phone']

        selected_institutional = [k for k in institutional_keys if k in selected_fields]
        selected_personal = [k for k in personal_keys if k in selected_fields]

        # Mapear key a etiqueta
        label_map = {
            'area': 'Dependencia',
            'cargo': 'Cargo',
            'remuneration': 'Remuneración',
            'blood_type': 'Tipo de Sangre',
            'marital_status': 'Estado Civil',
            'gender': 'Género',
            'birth_date': 'Fecha Nac.',
            'email': 'Correo Pers.',
            'address_reference': 'Dirección',
            'phone_number': 'Teléfono',
            'emergency_contact_name': 'Contacto Emerg.',
            'emergency_contact_phone': 'Celular Emerg.'
        }

        extra_labels = [label_map[k] for k in selected_institutional + selected_personal]
        headers = ['N°', 'Apellidos', 'Nombres', 'Cédula/Documento'] + extra_labels

        # Estilos para filas de unidad por profundidad (colores solicitados)
        depth_fills = [
            PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),  # azul padre
            PatternFill(start_color='198754', end_color='198754', fill_type='solid'),  # verde
            PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),  # amarillo
            PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),  # celeste
            PatternFill(start_color='FFD8A8', end_color='FFD8A8', fill_type='solid'),  # naranja claro
        ]

        # Contador de filas y índice global
        row_idx = 1

        # Añadimos encabezado general
        title_cell = ws.cell(row=row_idx, column=1, value=f"REPORTE: {report_scope_name}")
        title_cell.font = Font(bold=True, size=14)
        row_idx += 1

        # Escribir datos por unidad (cabecera por unidad con color y luego empleados)
        counter = 1
        for u, depth in units_with_depth:
            # Cabecera unidad
            col_count = len(headers)
            cell = ws.cell(row=row_idx, column=1, value=u.name)
            cell.font = Font(bold=True)
            fill = depth_fills[min(depth, len(depth_fills)-1)]
            for c in range(1, col_count + 1):
                ws.cell(row=row_idx, column=c).fill = fill
            row_idx += 1

            # Obtener empleados de la unidad (solo activos)
            employees = Employee.objects.filter(
                area_id=u.id,
                is_active=True
            ).exclude(
                Q(employment_status__name__icontains='EX EMPLEADO') |
                Q(employment_status__name__icontains='EX TRABAJADOR')
            ).select_related('person', 'area').order_by('person__last_name')

            # Si hay empleados, escribir cabecera de columnas antes del primer empleado
            if employees.exists():
                # First, write group header row (Institutional / Personal) if both present
                group_col_start = 5
                # institutional group length
                inst_len = len(selected_institutional)
                pers_len = len(selected_personal)

                if inst_len or pers_len:
                    # write empty cells for base columns
                    for bc in range(1, 5):
                        ws.cell(row=row_idx, column=bc, value='')
                    # merge and label institutional
                    col = group_col_start
                    if inst_len:
                        ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col+inst_len-1)
                        cell = ws.cell(row=row_idx, column=col, value='Datos Institucionales')
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(bold=True)
                    col += inst_len
                    if pers_len:
                        ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col+pers_len-1)
                        cell = ws.cell(row=row_idx, column=col, value='Datos Personales')
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(bold=True)
                    row_idx += 1

                # Header row with column titles
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    # add border
                    from openpyxl.styles import Border, Side
                    side = Side(border_style='thin', color='000000')
                    cell.border = Border(left=side, right=side, top=side, bottom=side)
                row_idx += 1

                for emp in employees:
                    last, first = emp.person.last_name if emp.person else '-', emp.person.first_name if emp.person else '-'
                    doc = emp.person.document_number if emp.person and emp.person.document_number else ''
                    values = [counter, last, first, doc]
                    for key in selected_institutional + selected_personal:
                        values.append(get_val(emp, key))
                    for col, val in enumerate(values, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        # add thin border to every cell
                        from openpyxl.styles import Border, Side
                        side = Side(border_style='thin', color='000000')
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                    row_idx += 1
                    counter += 1

        # Ajustar anchos
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18 if i > 1 else 6

        # Preparar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Reporte_Empleados_{report_scope_name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class PersonListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Person
    template_name = 'person/person_list.html'
    context_object_name = 'people'
    paginate_by = 10
    permission_required = 'person.view_person'

    def get_unit_tree_ids(self, unit_id):
        """Retorna la unidad seleccionada y todas sus subdependencias activas."""
        try:
            root_unit = AdministrativeUnit.objects.get(pk=unit_id, is_active=True)
        except AdministrativeUnit.DoesNotExist:
            return []

        unit_ids = [root_unit.id]
        frontier = [root_unit.id]

        while frontier:
            children_ids = list(
                AdministrativeUnit.objects.filter(
                    parent_id__in=frontier,
                    is_active=True
                ).values_list('id', flat=True)
            )
            if not children_ids:
                break
            unit_ids.extend(children_ids)
            frontier = children_ids

        return unit_ids

    def get_filtered_people_queryset(self):
        """Construye el queryset de personas con los filtros actuales de la solicitud."""
        qs = Person.objects.select_related(
            'document_type',
            'user',
            'employee_profile__area',
            'employee_profile__employment_status',
            'marital_status'
        ).annotate(
            full_name_str=Concat('first_name', Value(' '), 'last_name', output_field=CharField())
        ).order_by('-pk')

        q = self.request.GET.get('q')
        education_level_code = self.request.GET.get('education_level')
        cedula = self.request.GET.get('cedula')
        nombres = self.request.GET.get('nombres')
        area_id = self.request.GET.get('area')
        status_id = self.request.GET.get('status')
        marital_id = self.request.GET.get('marital_status')
        gender_id = self.request.GET.get('gender')

        has_advanced_search = any([cedula, nombres, area_id, status_id, marital_id, gender_id])

        if not has_advanced_search and not q:
            qs = qs.filter(employee_profile__is_active=True)

        if q:
            terms = [t.strip() for t in q.split() if t.strip()]
            for term in terms:
                qs = qs.filter(
                    Q(first_name__icontains=term) |
                    Q(last_name__icontains=term) |
                    Q(document_number__icontains=term) |
                    Q(email__icontains=term) |
                    Q(full_name_str__icontains=term)
                )

        if cedula:
            qs = qs.filter(document_number__icontains=cedula)

        if nombres:
            terms = [t.strip() for t in nombres.split() if t.strip()]
            for term in terms:
                qs = qs.filter(
                    Q(first_name__icontains=term) |
                    Q(last_name__icontains=term) |
                    Q(full_name_str__icontains=term)
                )

        if area_id:
            unit_ids = self.get_unit_tree_ids(area_id)
            if unit_ids:
                qs = qs.filter(employee_profile__area_id__in=unit_ids)
            else:
                qs = qs.filter(employee_profile__area_id=area_id)

        if education_level_code:
            codes = [c.strip() for c in education_level_code.split(',') if c.strip()]
            if len(codes) > 1:
                qs = qs.filter(curriculum__academic_titles__education_level__code__in=codes).distinct()
            else:
                qs = qs.filter(curriculum__academic_titles__education_level__code=codes[0]).distinct()

        if status_id:
            qs = qs.filter(employee_profile__employment_status_id=status_id)

        if marital_id:
            qs = qs.filter(marital_status_id=marital_id)

        if gender_id:
            qs = qs.filter(gender_id=gender_id)

        return qs

    def paginate_queryset(self, queryset, page_size):
        """Nunca lanza EmptyPage — corrige page fuera de rango."""
        paginator = self.get_paginator(queryset, page_size)
        try:
            page_num = int(self.request.GET.get('page', 1))
        except (ValueError, TypeError):
            page_num = 1
        # Clamp: nunca menor que 1, nunca mayor que el total
        page_num = max(1, min(page_num, paginator.num_pages or 1))
        page = paginator.page(page_num)
        return paginator, page, page.object_list, page.has_other_pages()

    def get_queryset(self):
        qs = self.get_filtered_people_queryset()

        # Aplicar ordenamiento personalizado desde parámetros GET (solo campos permitidos)
        sort_field = self.request.GET.get('sort_field')
        sort_dir = self.request.GET.get('sort_dir', 'asc')
        # Parámetros de orden enviados por el cliente
        allowed = {
            'full_name_str': 'last_name',
            'document_number': 'document_number',
            'email': 'email',
            'phone_number': 'phone_number',
            'employee_profile__area__name': 'employee_profile__area__name',
            'employee_profile__employment_status__name': 'employee_profile__employment_status__name'
        }
        if sort_field in allowed:
            field = allowed[sort_field]
            if sort_dir == 'desc':
                field = '-' + field
            qs = qs.order_by(field)
            # Orden aplicado correctamente (sin logs de depuración)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Formularios y Catálogos para el Modal de Búsqueda Avanzada
        context['form'] = PersonForm()
        context['marital_status_list'] = CatalogItem.objects.filter(catalog__code='MARITAL_STATUSES', is_active=True)
        context['areas_list'] = AdministrativeUnit.objects.filter(is_active=True)
        # Heredar todos los estados del catálogo EMPLOYMENT_STATUS
        context['status_list'] = CatalogItem.objects.filter(catalog__code='EMPLOYMENT_STATUS').order_by('name')
        context['gender_list'] = CatalogItem.objects.filter(catalog__code='GENDERS', is_active=True)

        # --- 3. ESTADÍSTICAS DINÁMICAS (con el filtro actual) ---
        from employee.models import Employee

        # Códigos de estados activos
        active_status_codes = ['EMPLEADO', 'TRABAJADOR', 'CONTRATADO', 'PROFESIONAL']

        filtered_people = getattr(self, 'object_list', None) or self.get_filtered_people_queryset()

        # Estadísticas sobre el resultado filtrado actual
        active_employees = Employee.objects.filter(
            person__in=filtered_people,
            employment_status__code__in=active_status_codes,
            is_active=True
        )

        stats_qs = active_employees.values(
            'employment_status__name',
            'employment_status__code',
            'employment_status__id'
        ).annotate(total=Count('id')).order_by('-total')

        # Convertimos a lista para fácil manejo en template
        stats = []
        total_active_employees = active_employees.count()

        # Tarjeta "Total Activos"
        stats.append({
            'label': 'Total Activos',
            'count': total_active_employees,
            'icon': 'fa-users',
            'class': 'color-one',
            'filter_val': ''  # Vacío limpia el filtro
        })

        # Tarjetas dinámicas por cada estado activo
        icons_map = {
            'EMPLEADO': 'fa-user-tie',
            'TRABAJADOR': 'fa-hard-hat',
            'CONTRATADO': 'fa-user-clock',
            'PROFESIONAL': 'fa-user-graduate'
        }
        colors = ['color-two', 'color-three', 'color-four', 'color-five']

        for idx, item in enumerate(stats_qs):
            code = item['employment_status__code']
            stats.append({
                'label': item['employment_status__name'] or 'Sin Estado',
                'count': item['total'],
                'icon': icons_map.get(code, 'fa-user-tag'),
                'class': colors[idx] if idx < len(colors) else 'color-one',
                'filter_val': item['employment_status__id']
            })


        context['stats_cards'] = stats
        return context

    def get(self, request, *args, **kwargs):
        # AJAX: devuelve solo la tabla parcial (búsqueda, paginación, filtros)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()

            # Modo exportación: devolver TODOS los registros sin paginación
            if request.GET.get('page_size') == '99999':
                from django.core.paginator import Paginator
                qs = self.object_list
                total = qs.count()
                # Un único "página" con todos los registros
                paginator = Paginator(qs, max(total, 1))
                page_obj = paginator.page(1)
                # Construir contexto mínimo sin pasar por paginate_queryset
                context = {
                    'people': qs,  # todos los registros
                    'page_obj': page_obj,
                    'paginator': paginator,
                    'is_paginated': False,
                    'request': request,
                }
            else:
                context = self.get_context_data()

            html = render_to_string('person/partials/partial_person_table.html', context, request=request)
            stats_html = render_to_string('person/partials/partial_person_stats.html', context, request=request)
            return JsonResponse({'success': True, 'html': html, 'stats_html': stats_html})

        return super().get(request, *args, **kwargs)


class PersonCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Person
    form_class = PersonForm
    template_name = 'person/modals/modal_person_form.html'
    permission_required = 'person.add_person'

    def has_permission(self):
        # Los superadministradores siempre tienen permiso
        if self.request.user.is_superuser:
            return True
        # Para otros usuarios, verificar permiso explícito
        return super().has_permission()

    def post(self, request, *args, **kwargs):
        # Nota: request.FILES es necesario para la foto
        try:
            form = PersonForm(request.POST, request.FILES)
            if form.is_valid():
                # Forzar que la persona creada quede activa
                person = form.save(commit=False)
                person.is_active = True
                person.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Persona registrada correctamente.',
                    # Devolvemos datos útiles por si quieres actualizar la tabla via JS sin recargar
                    'data': {'id': person.id, 'full_name': person.full_name}
                })
            # Loguear errores del formulario para facilitar debugging
            try:
                print("PersonCreateView - form invalid. POST keys:", dict(request.POST).keys())
                print("PersonCreateView - FILES keys:", request.FILES.keys())
                print("PersonCreateView - form errors:", form.errors.as_json())
            except Exception:
                pass
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
        except Exception as e:
            import traceback
            print("Error en PersonCreateView:", str(e))
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'Error interno: {str(e)}'}, status=500)


class PersonUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Person
    form_class = PersonForm
    template_name = 'person/modals/modal_person_form.html'
    permission_required = 'person.change_person'

    def has_permission(self):
        # Los superadministradores siempre tienen permiso
        if self.request.user.is_superuser:
            return True
        # Para otros usuarios, verificar permiso explícito
        return super().has_permission()

    def post(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()

            # Verificar si la foto actual existe físicamente
            if self.object.photo:
                try:
                    # Intentar acceder al archivo
                    if not self.object.photo.storage.exists(self.object.photo.name):
                        # La foto no existe físicamente, limpiar el campo
                        print(f"Foto no encontrada: {self.object.photo.name}, limpiando campo...")
                        self.object.photo = None
                        self.object.save(update_fields=['photo'])
                except Exception as e:
                    # Error al acceder a la foto (ruta inválida, etc), limpiar el campo
                    print(f"Error al verificar foto: {e}, limpiando campo...")
                    self.object.photo = None
                    self.object.save(update_fields=['photo'])

            # Procesar el formulario
            if 'photo' in request.FILES:
                # Hay una nueva foto
                form = PersonForm(request.POST, request.FILES, instance=self.object)
            else:
                # No hay nueva foto
                form = PersonForm(request.POST, instance=self.object)

            if form.is_valid():
                form.save()
                log_person_audit(
                    request,
                    self.object,
                    PersonAuditLog.Action.UPDATE,
                    PERSON_AUDIT_SECTIONS['personal']
                )
                return JsonResponse({'success': True, 'message': 'Datos actualizados correctamente.'})
            else:
                # Log de errores para debugging
                print("Errores de formulario:", form.errors)
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
        except Exception as e:
            # Log del error real
            import traceback
            print("Error en PersonUpdateView:", str(e))
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'Error interno: {str(e)}'}, status=500)


@method_decorator(require_POST, name='dispatch')
class PersonPhotoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'person.change_person'

    def post(self, request, *args, **kwargs):
        try:
            person = get_object_or_404(Person, pk=kwargs['pk'])
            photo = request.FILES.get('photo')

            if not photo:
                return JsonResponse({'success': False, 'message': 'Debe seleccionar una foto.'}, status=400)

            if hasattr(photo, 'size') and photo.size > 1 * 1024 * 1024:
                return JsonResponse({'success': False, 'message': 'La imagen es muy pesada. Máximo 1MB.'}, status=400)

            person.photo = photo
            person.save(update_fields=['photo'])
            log_person_audit(
                request,
                person,
                PersonAuditLog.Action.PHOTO,
                PERSON_AUDIT_SECTIONS['photo']
            )

            return JsonResponse({
                'success': True,
                'message': 'Foto actualizada correctamente.',
                'photo_url': person.photo.url if person.photo else None,
            })
        except Exception as e:
            import traceback
            print('Error en PersonPhotoUpdateView:', str(e))
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'Error interno: {str(e)}'}, status=500)


def person_detail_json(request, pk):
    p = get_object_or_404(Person, pk=pk)
    data = {
        'id': p.id,
        'document_type': p.document_type_id,
        'document_type_name': p.document_type.name if p.document_type else '',
        'document_number': p.document_number,
        'first_name': p.first_name,
        'last_name': p.last_name,
        'email': p.email,
        'birth_date': p.birth_date.isoformat() if p.birth_date else None,
        'gender': p.gender_id,
        'marital_status': p.marital_status_id,
        'blood_type': p.blood_type_id,
        'country': p.country_id,
        'province': p.province_id,
        'canton': p.canton_id,
        'parish': p.parish_id,
        'address_reference': p.address_reference,
        'phone_number': p.phone_number,
        'photo_url': p.photo.url if p.photo else None,
        # --- CAMPOS DE SALUD E INCLUSIÓN ---
        'has_disability': p.has_disability,
        'disability_type': p.disability_type_id,
        'disability_percentage': p.disability_percentage,
        'has_catastrophic_illness': p.has_catastrophic_illness,
        'catastrophic_illness_description': p.catastrophic_illness_description,
        'is_substitute': p.is_substitute,
        'substitute_family_member_id': p.substitute_family_member_id,
        'substitute_family_member_name': p.substitute_family_member_name,
        'substitute_family_member_relationship': p.substitute_family_member_relationship_id,
        'substitute_family_member_disability_type': p.substitute_family_member_disability_type_id,
        'substitute_family_member_disability_percentage': p.substitute_family_member_disability_percentage,
        # --- EMERGENCIA ---
        'emergency_contact_name': p.emergency_contact_name,
        'emergency_contact_phone': p.emergency_contact_phone,
        'emergency_contact_relationship': p.emergency_contact_relationship_id,
    }
    return JsonResponse({'success': True, 'data': data})


@login_required
@require_POST
def person_audit_log_api(request, pk):
    person = get_object_or_404(Person, pk=pk)
    action = (request.POST.get('action') or '').strip().upper()
    section = (request.POST.get('section') or '').strip()
    details = (request.POST.get('details') or '').strip()

    valid_actions = {choice[0] for choice in PersonAuditLog.Action.choices}
    if action not in valid_actions:
        return JsonResponse({'success': False, 'message': 'Acción inválida.'}, status=400)

    log = log_person_audit(request, person, action, section, details)
    return JsonResponse({
        'success': True,
        'message': 'Movimiento registrado correctamente.',
        'data': {
            'id': log.id,
            'movement_label': log.movement_label,
            'created_at': log.created_at.strftime('%d/%m/%Y %H:%M'),
        }
    })


@login_required
def person_audit_history_partial(request, pk):
    person = get_object_or_404(Person, pk=pk)
    query = (request.GET.get('q') or '').strip()
    page_number = request.GET.get('page') or 1
    export_excel = (request.GET.get('export') or '').strip() == '1'

    logs_queryset = PersonAuditLog.objects.filter(person=person).select_related('user')
    if query:
        logs_queryset = logs_queryset.filter(
            Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__username__icontains=query)
            | Q(section__icontains=query)
            | Q(details__icontains=query)
            | Q(action__icontains=query)
        )

    logs_queryset = logs_queryset.order_by('-created_at')

    if export_excel:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Auditoria Persona'
        worksheet.append(['Fecha y hora', 'Usuario', 'Movimiento', 'Sección', 'Detalle', 'IP'])

        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for log in logs_queryset:
            worksheet.append([
                log.created_at.strftime('%d/%m/%Y %H:%M'),
                getattr(log.user, 'get_full_name', lambda: '')() or log.user.username,
                log.get_action_display(),
                log.section or '',
                log.details or '',
                log.ip_address or '',
            ])

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value or '')))
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="auditoria_persona_{person.id}.xlsx"'
        workbook.save(response)
        return response

    paginator = Paginator(logs_queryset, 10)
    logs_page = paginator.get_page(page_number)
    html = render_to_string('person/partials/partial_person_audit_table.html', {
        'person': person,
        'audit_logs': logs_page.object_list,
        'page_obj': logs_page,
        'audit_query': query,
    }, request=request)
    return HttpResponse(html)


def person_quick_view_partial(request, pk):
    """
    Retorna un fragmento HTML con la información resumida de una persona.
    """
    person = get_object_or_404(
        Person.objects.select_related(
            'document_type',
            'gender',
            'employee_profile__area',
            'employee_profile__employment_status'
        ),
        pk=pk
    )
    # Recuperar objetos relacionados de forma segura.
    try:
        # Intentar usar el related_name si existe
        try:
            employee = person.employee_profile
        except Exception:
            employee = None

        budget = None
        institutional_data = None

        if employee:
            budget = BudgetLine.objects.filter(current_employee=employee).first()
            institutional_data = InstitutionalData.objects.filter(employee=employee).first()

        # Si faltan datos institucionales o partida, podemos mostrar el modal con mensaje informativo
        return render(request, 'person/partials/partial_person_quick_view.html', {
            'person': person,
            'budget': budget,
            'institutional_data': institutional_data
        })
    except Exception as e:
        # En caso de error inesperado, loguear y devolver un fragmento simple para el modal
        import traceback
        print('Error cargando vista rápida de persona:', str(e))
        traceback.print_exc()
        html = f"<div class='p-4 text-error'>No se pudo cargar la información completa: {str(e)}</div>"
        return HttpResponse(html, status=200)


@method_decorator(require_POST, name='dispatch')
class RelocateEmployeeView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'person.change_person'  # Ajusta según tus permisos

    def post(self, request, *args, **kwargs):
        person_id = request.POST.get('person_id')
        unit_id = request.POST.get('unit_id')

        if not person_id or not unit_id:
            return JsonResponse({'success': False, 'message': 'Faltan datos (Persona o Unidad).'})

        try:
            # 1. Obtener la persona
            person = get_object_or_404(Person, pk=person_id)

            # 2. Verificar que tenga perfil de empleado
            if not hasattr(person, 'employee_profile'):
                return JsonResponse({'success': False, 'message': 'Esta persona no tiene perfil de empleado activo.'})

            # 3. Obtener la unidad destino
            new_unit = get_object_or_404(AdministrativeUnit, pk=unit_id)

            # 4. Actualizar
            employee = person.employee_profile
            old_unit_name = employee.area.name if employee.area else "Sin Unidad"

            employee.area = new_unit
            employee.save()

            return JsonResponse({
                'success': True,
                'message': f'Reubicación exitosa: {person.first_name}  {person.last_name} pasó de "{old_unit_name}" a "{new_unit.name}".'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
