# apps/function_manual/views.py
import openpyxl
from django.db import models, transaction
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.views.generic import ListView, CreateView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import permission_required, login_required
import json

from openpyxl.styles import Border, PatternFill, Font, Alignment, Side

from institution.models import Deliverable as InstDeliverable
from employee.models import Employee
from institution.models import AdministrativeUnit
from .models import Competency, JobProfile, ManualCatalog, OccupationalMatrix, ManualCatalogItem, ValuationNode, \
    JobActivity, ProfileCompetency
from .forms import ManualCatalogForm, ManualCatalogItemForm
from core.models import BaseModel, Authorities


# ============================================================================
# 1. MIXINS Y AUXILIARES
# ============================================================================

class JobProfileMixin:
    """Mixin centralizado para proveer catálogos a cualquier vista del manual"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        def get_items(catalog_code):
            items = ManualCatalogItem.objects.filter(
                catalog__code=catalog_code, is_active=True
            ).select_related('target_role')
            return [
                {
                    'id': item.id,
                    'name': item.name,
                    'target_role': item.target_role_id if item.target_role else None
                }
                for item in items
            ]

        matrix_data = list(OccupationalMatrix.objects.all().values(
            'id', 'occupational_group', 'grade', 'remuneration',
            'required_role_id', 'minimum_instruction_id',
            'minimum_experience_months', 'required_decision_id',
            'required_impact_id', 'complexity_level_id'
        ))

        catalogs_dict = {
            "instruction": get_items('INSTRUCTION_LEVELS'),
            "decisions": get_items('DECISION_LEVELS'),
            "impact": get_items('IMPACT_LEVELS'),
            "roles": get_items('JOB_ROLES'),
            "verbs": get_items('ACTION_VERBS'),
            "frequency": get_items('FREQUENCY'),
            "complexity": get_items('COMPLEXITY_LEVELS'),
            "matrix": matrix_data,
            # Agregamos todas las competencias para filtrar en Vue
            "competencies": list(Competency.objects.filter(is_active=True).values('id', 'name', 'type', 'definition'))
        }
        context['catalogs_json'] = json.dumps(catalogs_dict, cls=DjangoJSONEncoder)
        return context


def get_manual_catalog_stats():
    """Retorna estadísticas de catálogos"""
    total = ManualCatalog.objects.count()
    active = ManualCatalog.objects.filter(is_active=True).count()
    return {'total': total, 'active': active, 'inactive': total - active}


# ============================================================================
# 2. VISTAS DE PERFILES (WIZARD)
# ============================================================================

class JobProfileListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = JobProfile
    permission_required = "function_manual.view_jobprofile"
    context_object_name = "profiles"

    def get_template_names(self):
        if self.request.GET.get('partial'):
            return ["function_manual/partials/partial_function_manual_table.html"]
        return ["function_manual/function_manual_list.html"]

    def get_queryset(self):
        queryset = JobProfile.objects.select_related(
            'administrative_unit',
            'occupational_classification',
            'referential_employee__person'  # Optimización para mostrar el empleado referencial
        ).filter(is_active=True)

        # Filtrar por usuario: solo administradores ven todos los perfiles
        user = self.request.user
        can_view_all = (
                user.has_perm('function_manual.can_admin') or
                user.is_superuser
        )

        if not can_view_all:
            # Usuarios normales solo ven sus propios perfiles
            queryset = queryset.filter(created_by=user)

        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                models.Q(specific_job_title__icontains=query) | models.Q(administrative_unit__name__icontains=query))
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not self.request.GET.get('partial'):
            qs = self.get_queryset()
            context['stats_total'] = qs.count()
            context['stats_classified'] = qs.filter(occupational_classification__isnull=False).count()
            context['stats_pending'] = qs.filter(occupational_classification__isnull=True).count()
            context['stats_active'] = qs.filter(is_active=True).count()

            # Contexto para el modal de asignación de grupo
            # Serializamos a JSON para manejo dinámico en Vue
            matrix_data = OccupationalMatrix.objects.select_related('required_role').values(
                'id', 'occupational_group', 'grade', 'remuneration', 'complexity_level_id'
            ).order_by('grade')
            context['occupational_matrix_json'] = json.dumps(list(matrix_data), cls=DjangoJSONEncoder)

        return context


class JobProfileCreateView(LoginRequiredMixin, PermissionRequiredMixin, JobProfileMixin, CreateView):
    model = JobProfile
    template_name = "function_manual/function_manual_form.html"
    permission_required = "function_manual.add_jobprofile"
    fields = ['position_code', 'specific_job_title', 'administrative_unit', 'referential_employee']
    success_url = reverse_lazy('function_manual:profile_list')


class JobProfileUpdateView(LoginRequiredMixin, PermissionRequiredMixin, JobProfileMixin, UpdateView):
    model = JobProfile
    template_name = "function_manual/function_manual_form.html"
    permission_required = "function_manual.change_jobprofile"
    fields = ['position_code', 'specific_job_title', 'administrative_unit', 'referential_employee']
    success_url = reverse_lazy('function_manual:profile_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.object

        # 1. Reconstruir ruta organizacional
        units_path = []
        current_unit = profile.administrative_unit
        while current_unit:
            units_path.insert(0, current_unit.id)
            current_unit = current_unit.parent

        # 2. Reconstruir ruta de valoración (Valuation Nodes)
        # Intentar reconstruir desde los campos individuales del perfil
        nodes_path = []

        # Si el perfil tiene clasificación ocupacional, usar esa ruta completa
        if profile.occupational_classification:
            leaf_node = ValuationNode.objects.filter(
                occupational_classification=profile.occupational_classification,
                node_type='RESULT'
            ).first()

            if leaf_node:
                curr = leaf_node
                while curr:
                    nodes_path.insert(0, curr.id)
                    curr = curr.parent
        else:
            # Si no tiene clasificación, reconstruir desde los campos individuales
            # Buscar cada nodo basándose en los catalog_items guardados

            # 1. ROL
            if profile.job_role:
                role_node = ValuationNode.objects.filter(
                    node_type='ROLE',
                    catalog_item=profile.job_role,
                    is_active=True
                ).first()
                if role_node:
                    nodes_path.append(role_node.id)

                    # 2. INSTRUCCIÓN
                    if profile.required_instruction:
                        instruction_node = ValuationNode.objects.filter(
                            node_type='INSTRUCTION',
                            catalog_item=profile.required_instruction,
                            parent=role_node,
                            is_active=True
                        ).first()
                        if instruction_node:
                            nodes_path.append(instruction_node.id)

                            # 3. EXPERIENCIA (buscar el primer nodo hijo)
                            experience_node = ValuationNode.objects.filter(
                                node_type='EXPERIENCE',
                                parent=instruction_node,
                                is_active=True
                            ).first()
                            if experience_node:
                                nodes_path.append(experience_node.id)

                                # 4. DECISIÓN
                                if profile.decision_making:
                                    decision_node = ValuationNode.objects.filter(
                                        node_type='DECISION',
                                        catalog_item=profile.decision_making,
                                        parent=experience_node,
                                        is_active=True
                                    ).first()
                                    if decision_node:
                                        nodes_path.append(decision_node.id)

                                        # 5. IMPACTO
                                        if profile.management_impact:
                                            impact_node = ValuationNode.objects.filter(
                                                node_type='IMPACT',
                                                catalog_item=profile.management_impact,
                                                parent=decision_node,
                                                is_active=True
                                            ).first()
                                            if impact_node:
                                                nodes_path.append(impact_node.id)

                                                # 6. COMPLEJIDAD
                                                if profile.final_complexity_level:
                                                    complexity_node = ValuationNode.objects.filter(
                                                        node_type='COMPLEXITY',
                                                        catalog_item=profile.final_complexity_level,
                                                        parent=impact_node,
                                                        is_active=True
                                                    ).first()
                                                    if complexity_node:
                                                        nodes_path.append(complexity_node.id)

        # 3. Actividades
        activities_data = []
        for act in profile.activities.all():
            activities_data.append({
                'action_verb': act.action_verb_id,
                'description': act.description,
                'additional_knowledge': act.additional_knowledge,
                'deliverable': act.deliverable_id,
                'complexity': act.complexity_id,
                'contribution': act.contribution_id,
                'frequency': act.frequency_id,
            })
        unit_deliverables = list(InstDeliverable.objects.filter(
            unit=profile.administrative_unit, is_active=True
        ).values('id', 'name'))
        context['unit_deliverables_json'] = json.dumps(unit_deliverables, cls=DjangoJSONEncoder)
        # 4. Competencias
        comp_map = {'TECHNICAL': [], 'BEHAVIORAL': [], 'TRANSVERSAL': []}
        # Ordenamos por id o algo consistente para que aparezcan en los slots correctos
        for pc in profile.profilecompetency_set.select_related('competency').order_by('id'):
            if pc.competency.type in comp_map:
                comp_map[pc.competency.type].append(pc.competency.id)

        # Rellenar con strings vacíos para satisfacer los arrays de tamaño fijo del front (3, 3, 2)
        while len(comp_map['TECHNICAL']) < 3: comp_map['TECHNICAL'].append('')
        while len(comp_map['BEHAVIORAL']) < 3: comp_map['BEHAVIORAL'].append('')
        while len(comp_map['TRANSVERSAL']) < 2: comp_map['TRANSVERSAL'].append('')

        # 5. Payload Final
        initial_data = {
            'id': profile.id,
            'position_code': profile.position_code,
            'specific_job_title': profile.specific_job_title,
            'mission': profile.mission,
            'knowledge_area': profile.knowledge_area,
            'experience_details': profile.experience_details,
            'training_topic': profile.training_topic or '',
            'interface_relations': profile.interface_relations or '',
            'selectedUnits': units_path,
            # Importante: para que el wizard vue cargue paso a paso, puede requerir que selectedNodes
            # se reconstruya secuencialmente. Pero probemos pasando el array completo.
            'selectedNodes': nodes_path,
            'activities': activities_data,
            'selectedTechnical': comp_map['TECHNICAL'][:3],
            'selectedBehavioral': comp_map['BEHAVIORAL'][:3],
            'selectedTransversal': comp_map['TRANSVERSAL'][:2],
            'matchResult': {
                'group': profile.occupational_classification.occupational_group,
                'grade': profile.occupational_classification.grade,
                'remuneration': str(profile.occupational_classification.remuneration),
                'id': profile.occupational_classification.id
            } if profile.occupational_classification else None
        }

        context['initial_data'] = json.dumps(initial_data, cls=DjangoJSONEncoder)
        return context


class JobProfileAssignReferentialView(LoginRequiredMixin, View):
    """
    Vista AJAX para asignar un empleado referencial a un perfil ya creado.
    """

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        return render(request, 'function_manual/modals/modal_assign_referential.html', {'profile': profile})

    def post(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        employee_id = request.POST.get('employee_id')
        try:
            emp = Employee.objects.get(pk=employee_id)
            profile.referential_employee = emp
            profile.save()
            return JsonResponse({'success': True, 'message': 'Empleado referencial asignado con éxito.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
def api_get_available_roles(request):
    """Obtener todos los nodos de tipo ROLE activos con sus nombres legibles"""
    try:
        roles = ValuationNode.objects.filter(
            node_type='ROLE',
            is_active=True
        ).select_related('catalog_item').order_by('catalog_item__name')

        roles_data = [
            {
                'id': role.id,
                'name': role.catalog_item.name if role.catalog_item else f"Rol #{role.id}"
            }
            for role in roles
        ]

        return JsonResponse({'success': True, 'data': roles_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_search_employee_simple(request):
    """Búsqueda simple de empleado por cédula (sin bloqueo por partida)"""
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'success': False, 'message': 'Ingrese cédula'})

    try:
        emp = Employee.objects.select_related('person').get(person__document_number=q, is_active=True)

        # Check photo URL safely
        photo_url = None
        if emp.person.photo:
            try:
                photo_url = emp.person.photo.url
            except Exception:
                photo_url = None

        return JsonResponse({
            'success': True,
            'data': {
                'id': emp.id,
                'full_name': str(emp.person),
                'photo': photo_url,
                'email': emp.person.email or 'Sin email'
            }
        })
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado o inactivo'})
    except Employee.MultipleObjectsReturned:
        return JsonResponse({'success': False, 'message': 'Error de datos: Cédula duplicada en sistema'})
    except Exception as e:
        import traceback
        print(traceback.format_exc())  # Print to server console too
        return JsonResponse({'success': False, 'message': f'Error del servidor: {str(e)}'})


# ============================================================================
# 3. ESCALAS / ESTRUCTURA DE VALORACIÓN (DRILL-DOWN)
# ============================================================================

class OccupationalMatrixListView(LoginRequiredMixin, PermissionRequiredMixin, JobProfileMixin, ListView):
    """Listado plano de la escala salarial (SP)"""
    model = OccupationalMatrix
    template_name = "function_manual/occupational_matrix_list.html"
    context_object_name = "matrix_entries"
    permission_required = "function_manual.view_occupationalmatrix"

    def get_queryset(self):
        return OccupationalMatrix.objects.select_related(
            'required_role', 'complexity_level', 'minimum_instruction'
        ).all().order_by('grade')


class OccupationalMatrixSaveApi(LoginRequiredMixin, View):
    """API para crear o editar un grado salarial"""

    def post(self, request):
        data = json.loads(request.body)
        entry_id = data.get('id')

        if entry_id:
            entry = get_object_or_404(OccupationalMatrix, pk=entry_id)
        else:
            entry = OccupationalMatrix()

        entry.occupational_group = data.get('occupational_group')
        entry.grade = data.get('grade')
        entry.remuneration = data.get('remuneration')
        entry.required_role_id = data.get('required_role_id')
        entry.complexity_level_id = data.get('complexity_level_id')
        entry.minimum_instruction_id = data.get('minimum_instruction_id')
        entry.minimum_experience_months = data.get('minimum_experience_months')
        entry.save()

        return JsonResponse({'success': True, 'message': 'Escala salarial actualizada.'})


# ============================================================================
# 4. GESTIÓN DE CATÁLOGOS (La sección que faltaba)
# ============================================================================

class ManualCatalogListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ManualCatalog
    template_name = 'function_manual/catalog_list.html'
    context_object_name = 'catalogs'
    permission_required = 'function_manual.view_manualcatalog'

    def get_queryset(self):
        query = self.request.GET.get('q')
        qs = ManualCatalog.objects.all()
        if query: qs = qs.filter(name__icontains=query)
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ManualCatalogForm()
        context['item_form'] = ManualCatalogItemForm()  # Agregamos el formulario de items
        stats = get_manual_catalog_stats()
        context.update(
            {'stats_total': stats['total'], 'stats_active': stats['active'], 'stats_inactive': stats['inactive']})
        return context


class ManualCatalogCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ManualCatalog
    form_class = ManualCatalogForm
    permission_required = 'function_manual.add_manualcatalog'

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            catalog = form.save()
            return JsonResponse({'success': True, 'message': 'Catálogo creado.'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


class ManualCatalogUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ManualCatalog
    form_class = ManualCatalogForm
    permission_required = 'function_manual.change_manualcatalog'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Catálogo actualizado.'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


def manual_catalog_detail_json(request, pk):
    c = get_object_or_404(ManualCatalog, pk=pk)
    return JsonResponse({'success': True,
                         'data': {'id': c.id, 'name': c.name, 'code': c.code, 'description': c.description or '',
                                  'is_active': c.is_active}})


@require_POST
def manual_catalog_toggle_status(request, pk):
    c = get_object_or_404(ManualCatalog, pk=pk)
    c.toggle_status()
    return JsonResponse({'success': True})


# --- VISTAS DE ITEMS DE CATÁLOGO ---

class ManualCatalogItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ManualCatalogItem
    form_class = ManualCatalogItemForm
    permission_required = 'function_manual.add_manualcatalogitem'

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            item = form.save(commit=False)
            item.catalog_id = request.POST.get('catalog_id')
            item.save()
            return JsonResponse({'success': True, 'message': 'Item creado.'})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


class ManualCatalogItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ManualCatalogItem
    form_class = ManualCatalogItemForm
    permission_required = 'function_manual.change_manualcatalogitem'

    def post(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
            form = self.get_form()
            if form.is_valid():
                form.save()
                return JsonResponse({'success': True, 'message': 'Item actualizado.'})
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Error interno: {str(e)}"}, status=500)


def manual_catalog_item_list_json(request, catalog_id):
    items = ManualCatalogItem.objects.filter(catalog_id=catalog_id).values('id', 'code', 'name', 'description',
                                                                           'is_active')
    return JsonResponse({'success': True, 'data': list(items)})


def manual_catalog_item_detail_json(request, pk):
    i = get_object_or_404(ManualCatalogItem, pk=pk)
    return JsonResponse({'success': True,
                         'data': {'id': i.id, 'catalog_id': i.catalog_id,
                                  'catalog_name': i.catalog.name,  # Needed for JS
                                  'name': i.name, 'code': i.code,
                                  'description': i.description or '',
                                  'target_role': i.target_role_id if i.target_role else None}})


@require_POST
def manual_catalog_item_toggle_status(request, pk):
    i = get_object_or_404(ManualCatalogItem, pk=pk)
    i.toggle_status()
    return JsonResponse({'success': True})


# ============================================================================
# 5. COMPETENCIAS
# ============================================================================

class CompetencyListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Competency
    template_name = "function_manual/competency_list.html"
    permission_required = "function_manual.view_competency"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        complexity_items = ManualCatalogItem.objects.filter(catalog__code='COMPLEXITY_LEVELS', is_active=True)
        context['complexity_levels'] = json.dumps(
            [{'id': item.id, 'name': item.name, 'code': item.code} for item in complexity_items])
        qs = Competency.objects.filter(is_active=True)
        context.update({'stats_total': qs.count(), 'stats_behavioral': qs.filter(type='BEHAVIORAL').count(),
                        'stats_technical': qs.filter(type='TECHNICAL').count(),
                        'stats_transversal': qs.filter(type='TRANSVERSAL').count()})
        return context


class CompetencyTablePartialView(CompetencyListView):
    template_name = "function_manual/partials/partial_competency_table.html"


class CompetencyCreateView(LoginRequiredMixin, View):
    def post(self, request):
        data = json.loads(request.body)
        Competency.objects.create(name=data.get('name'), type=data.get('type'), definition=data.get('definition'),
                                  suggested_level_id=data.get('suggested_level') or None)
        return JsonResponse({'status': 'success', 'message': 'Competencia creada.'})


class CompetencyUpdateView(LoginRequiredMixin, UpdateView):
    model = Competency

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        c = self.get_object()
        c.name, c.type, c.definition, c.suggested_level_id = data.get('name'), data.get('type'), data.get(
            'definition'), data.get('suggested_level') or None
        c.save()
        return JsonResponse({'status': 'success', 'message': 'Competencia actualizada.'})


class CompetencyToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        c = get_object_or_404(Competency, pk=pk)
        c.is_active = not c.is_active
        c.save()
        return JsonResponse({'status': 'success'})


# ============================================================================
# 6. APIs DE SOPORTE (Navegación y Autocompletado)
# ============================================================================

class ApiValuationNodesView(View):
    def get(self, request):
        parent_id = request.GET.get('parent')
        filters = {'parent_id': parent_id} if parent_id else {'parent__isnull': True}
        nodes = ValuationNode.objects.filter(**filters, is_active=True).select_related(
            'catalog_item', 'occupational_classification'
        )

        def get_node_name(n):
            if n.catalog_item:
                return n.catalog_item.name
            if n.node_type == 'RESULT' and n.occupational_classification:
                return f"{n.occupational_classification.occupational_group} - G{n.occupational_classification.grade}"
            return n.name_extra if n.name_extra else "Sin Definición"

        data = [{'id': n.id, 'name': get_node_name(n), 'type': n.node_type,
                 'catalog_item_id': n.catalog_item_id,
                 'classification_id': n.occupational_classification_id,
                 'classification': {'group': n.occupational_classification.occupational_group,
                                    'rmu': n.occupational_classification.remuneration,
                                    'grade': n.occupational_classification.grade} if n.node_type == 'RESULT' and n.occupational_classification else None}
                for n in nodes]
        return JsonResponse(data, safe=False)


class ValuationNodeDetailApi(View):
    def get(self, request, pk):
        n = get_object_or_404(ValuationNode, pk=pk)
        return JsonResponse(
            {'id': n.id, 'node_type': n.node_type, 'catalog_item_id': n.catalog_item_id, 'name_extra': n.name_extra,
             'occupational_classification_id': n.occupational_classification_id})


class ValuationNodeSaveApi(LoginRequiredMixin, View):
    def post(self, request):
        data = json.loads(request.body)
        node_id = data.get('id')
        node = get_object_or_404(ValuationNode, pk=node_id) if node_id else ValuationNode(
            parent_id=data.get('parent_id'), node_type=data.get('node_type'))
        node.catalog_item_id, node.name_extra, node.occupational_classification_id = data.get(
            'catalog_item_id') or None, data.get('name_extra'), data.get('occupational_classification_id') or None
        node.save()
        return JsonResponse({'success': True})


class ApiNextPositionCodeView(View):
    def get(self, request, unit_id):
        unit = get_object_or_404(AdministrativeUnit, pk=unit_id)
        last = JobProfile.objects.filter(administrative_unit_id=unit_id).order_by('position_code').last()
        next_num = (int(last.position_code.split('.')[-1]) + 1) if (last and last.position_code) else 1
        return JsonResponse({'next_code': f"{unit.code or '0'}.{next_num:02d}"})


class ApiUnitChildrenView(View):
    def get(self, request, parent_id=None):
        filters = {'parent_id': parent_id} if parent_id else {'parent__isnull': True}
        units = AdministrativeUnit.objects.filter(**filters, is_active=True).values('id', 'name', 'code')
        return JsonResponse(list(units), safe=False)


class ValuationNodeListView(LoginRequiredMixin, PermissionRequiredMixin, JobProfileMixin, ListView):
    """
    Gestiona la estructura jerárquica de valoración (Drill-down).
    Esta es la vista que el urls.py busca como 'valuation_list'.
    """
    model = ValuationNode
    template_name = "function_manual/valuation_structure_list.html"
    context_object_name = "nodes"
    permission_required = "function_manual.view_valuationnode"

    def get_queryset(self):
        parent_id = self.request.GET.get('parent')
        if parent_id:
            return ValuationNode.objects.filter(parent_id=parent_id, is_active=True)
        # Nivel 1: ROLES (Raíz)
        return ValuationNode.objects.filter(parent__isnull=True, is_active=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        parent_id = self.request.GET.get('parent')

        # Valores iniciales (Raíz)
        context['parent_name'] = "Estructura Raíz"
        context['next_level_name'] = "Rol de Puesto"
        context['next_node_type'] = "ROLE"
        context['breadcrumbs'] = []

        if parent_id:
            parent = get_object_or_404(ValuationNode, pk=parent_id)
            context['parent_node'] = parent
            context['parent_name'] = parent.catalog_item.name if parent.catalog_item else parent.name_extra

            # Reconstruir camino de Breadcrumbs
            bc = []
            curr = parent
            while curr:
                bc.insert(0, curr)
                curr = curr.parent
            context['breadcrumbs'] = bc

            # Lógica de la Norma: Mapeo de tipos para el siguiente nivel
            mapping = {
                'ROLE': ('INSTRUCTION', 'Instrucción Formal'),
                'INSTRUCTION': ('EXPERIENCE', 'Experiencia Requerida'),
                'EXPERIENCE': ('DECISION', 'Nivel de Decisiones'),
                'DECISION': ('IMPACT', 'Nivel de Impacto'),
                'IMPACT': ('COMPLEXITY', 'Nivel de Complejidad'),
                'COMPLEXITY': ('RESULT', 'Resultado Final')
            }
            res = mapping.get(parent.node_type, ('RESULT', 'Resultado'))
            context['next_node_type'] = res[0]
            context['next_level_name'] = res[1]

        return context


class OccupationalMatrixDetailApi(View):
    """Retorna el JSON de un grado salarial para editar"""

    def get(self, request, pk):
        entry = get_object_or_404(OccupationalMatrix, pk=pk)
        return JsonResponse({
            'id': entry.id,
            'occupational_group': entry.occupational_group,
            'grade': entry.grade,
            'remuneration': entry.remuneration,
            'minimum_experience_months': entry.minimum_experience_months,
            'required_role_id': entry.required_role_id,
            'minimum_instruction_id': entry.minimum_instruction_id,
            'complexity_level_id': entry.complexity_level_id,
        })


@require_POST
def occupational_matrix_toggle_status(request, pk):
    """Da de baja o activa un grado salarial"""
    entry = get_object_or_404(OccupationalMatrix, pk=pk)
    entry.is_active = not entry.is_active
    entry.save()
    return JsonResponse({'success': True, 'is_active': entry.is_active})


class JobProfileSaveApi(LoginRequiredMixin, View):
    """
    Procesa el guardado masivo del Perfil y sus Actividades en una sola transacción.
    """

    @method_decorator(csrf_protect)
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            with transaction.atomic():
                profile_id = data.get('id')
                # 1. Crear o Actualizar Perfil
                if profile_id:
                    profile = JobProfile.objects.get(pk=profile_id)
                else:
                    profile = JobProfile()

                profile.position_code = data.get('position_code')
                profile.specific_job_title = data.get('specific_job_title')
                profile.administrative_unit_id = data.get('administrative_unit')

                # Guardar todos los niveles de valoración desde los nodos seleccionados
                role_node_id = data.get('role_node_id')
                if role_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=role_node_id)
                        if node.catalog_item:
                            profile.job_role = node.catalog_item
                    except ValuationNode.DoesNotExist:
                        pass

                instruction_node_id = data.get('instruction_node_id')
                if instruction_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=instruction_node_id)
                        if node.catalog_item:
                            profile.required_instruction = node.catalog_item
                    except ValuationNode.DoesNotExist:
                        pass

                experience_node_id = data.get('experience_node_id')
                if experience_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=experience_node_id)
                        if node.name_extra:
                            # Extraer meses de experiencia del nombre
                            import re
                            match = re.search(r'(\d+)', node.name_extra)
                            if match:
                                profile.required_experience_months = int(match.group(1))
                            elif 'no requerida' in node.name_extra.lower():
                                profile.required_experience_months = 0
                    except ValuationNode.DoesNotExist:
                        pass

                decision_node_id = data.get('decision_node_id')
                if decision_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=decision_node_id)
                        if node.catalog_item:
                            profile.decision_making = node.catalog_item
                    except ValuationNode.DoesNotExist:
                        pass

                impact_node_id = data.get('impact_node_id')
                if impact_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=impact_node_id)
                        if node.catalog_item:
                            profile.management_impact = node.catalog_item
                    except ValuationNode.DoesNotExist:
                        pass

                complexity_node_id = data.get('complexity_node_id')
                if complexity_node_id:
                    try:
                        node = ValuationNode.objects.get(pk=complexity_node_id)
                        if node.catalog_item:
                            profile.final_complexity_level = node.catalog_item
                    except ValuationNode.DoesNotExist:
                        pass

                # Retrieve Matrix details
                matrix_id = data.get('occupational_classification')
                profile.occupational_classification_id = matrix_id

                if matrix_id:
                    matrix = OccupationalMatrix.objects.get(pk=matrix_id)
                    profile.job_role = matrix.required_role
                    profile.required_instruction = matrix.minimum_instruction
                    profile.decision_making = matrix.required_decision
                    profile.management_impact = matrix.required_impact
                    profile.final_complexity_level = matrix.complexity_level
                    profile.required_experience_months = matrix.minimum_experience_months

                profile.mission = data.get('mission')
                profile.knowledge_area = data.get('knowledge_area')
                profile.experience_details = data.get('experience_details')
                profile.training_topic = data.get('training_topic')
                profile.interface_relations = data.get('interface_relations')
                profile.is_active = True
                profile.save()

                # 2. Gestionar Actividades (Borrado y re-inserción para simplicidad en Wizard)
                profile.activities.all().delete()
                activities_data = data.get('activities', [])

                activities_to_create = []
                for act in activities_data:
                    if act.get('description') and act.get('action_verb'):
                        activities_to_create.append(JobActivity(
                            profile=profile,
                            action_verb_id=act.get('action_verb'),
                            description=act.get('description'),
                            additional_knowledge=act.get('additional_knowledge', ''),
                            deliverable_id=act.get('deliverable') or None,
                            complexity_id=act.get('complexity') or None,
                            contribution_id=act.get('contribution') or None,
                            frequency_id=act.get('frequency') or None,
                        ))

                if activities_to_create:
                    JobActivity.objects.bulk_create(activities_to_create)
                else:
                    raise ValueError("El perfil debe tener al menos una actividad esencial.")

                # 3. Gestionar Competencias (Borrado y re-inserción)
                # data.get('competencies') debe ser un array de IDs
                profile.competencies.clear()
                comp_ids = data.get('competencies', [])
                # Filtramos IDs válidos y únicos
                comp_ids = list(set([cid for cid in comp_ids if cid]))

                # Para masivo con observable_behavior si fuera necesario, aquí usamos through default o create
                # Como profile.competencies es m2m through ProfileCompetency, usamos el manager directo si el through permite o bulk_create del through
                # Dado que ProfileCompetency tiene observable_behavior opcional ahora:

                comps_to_create = []
                for cid in comp_ids:
                    # Buscamos la definición para pre-llenar observable_behavior si se desea, por ahora vacío
                    comps_to_create.append(ProfileCompetency(
                        profile=profile,
                        competency_id=cid,
                        observable_behavior=''  # Opcional
                    ))
                if comps_to_create:
                    ProfileCompetency.objects.bulk_create(comps_to_create)

                return JsonResponse({
                    'success': True,
                    'message': 'Perfil de Puesto guardado correctamente.',
                    'redirect': reverse_lazy('function_manual:profile_list')
                })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)


class JobProfileLegalizeView(LoginRequiredMixin, View):
    """Vista para legalizar (firmas) el perfil de puesto"""

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        authorities = Authorities.objects.filter(is_active=True)
        return render(request, 'function_manual/modals/modal_legalize_profile.html', {
            'profile': profile,
            'authorities': authorities
        })

    def post(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        try:
            profile.prepared_by_id = request.POST.get('prepared_by') or None
            profile.reviewed_by_id = request.POST.get('reviewed_by') or None
            profile.approved_by_id = request.POST.get('approved_by') or None
            profile.save()

            return JsonResponse({'success': True, 'message': 'Firmas de legalización actualizadas correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)


class JobProfileUploadLegalizedView(LoginRequiredMixin, View):
    """Sube el archivo PDF legalizado"""

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        return render(request, 'function_manual/modals/modal_upload_legalized.html', {'profile': profile})

    def post(self, request, pk):
        profile = get_object_or_404(JobProfile, pk=pk)
        if 'legalized_document' in request.FILES:
            profile.legalized_document = request.FILES['legalized_document']
            profile.save()
            return JsonResponse({'success': True, 'message': 'Archivo subido correctamente.'})
        return JsonResponse({'success': False, 'message': 'No se recibió ningún archivo.'}, status=400)


class JobProfileDetailModalView(LoginRequiredMixin, View):
    """Muestra el detalle del perfil y opción para imprimir"""

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile.objects.select_related(
            'administrative_unit', 'occupational_classification', 'referential_employee__person',
            'prepared_by', 'reviewed_by', 'approved_by'
        ).prefetch_related('activities', 'competencies'), pk=pk)

        return render(request, 'function_manual/modals/modal_profile_detail.html', {'profile': profile})


class JobProfilePrintView(LoginRequiredMixin, View):
    """Vista para generar la impresión del perfil"""

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile.objects.select_related(
            'administrative_unit', 'occupational_classification',
            'required_instruction', 'job_role',
            'prepared_by',
            'reviewed_by',
            'approved_by'
        ), pk=pk)

        activities = list(profile.activities.all())

        # Separar competencias por tipo
        competencies_qs = ProfileCompetency.objects.filter(profile=profile).select_related('competency',
                                                                                           'competency__suggested_level')
        technical = [c for c in competencies_qs if c.competency.type == 'TECHNICAL']
        behavioral = [c for c in competencies_qs if c.competency.type == 'BEHAVIORAL']
        transversal = [c for c in competencies_qs if c.competency.type == 'TRANSVERSAL']

        # Rango para llenar filas vacías en actividades si son pocas (estética)
        empty_activities_range = range(5 - len(activities)) if len(activities) < 5 else []

        context = {
            'profile': profile,
            'activities': activities,
            'technical_competencies': technical,
            'behavioral_competencies': behavioral,
            'transversal_competencies': transversal,
            'empty_activities_range': empty_activities_range
        }
        return render(request, 'function_manual/print_profile.html', context)


class JobProfileValuationExcelView(LoginRequiredMixin, View):
    """Genera reporte Excel de Valoración de Puestos (Legalizados) - Estilo mdT"""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from django.utils import timezone

        # 1. Configuración de Respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Reporte_Valoracion_Puestos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # 2. Crear Libro y Hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valoración de Puestos"

        # 3. Estilos
        # Bordes
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Fuentes
        font_header_main = Font(name='Arial', size=11, bold=True, color='000000')
        font_header_sub = Font(name='Arial', size=10, bold=True)
        font_body = Font(name='Arial', size=9)

        # Rellenos (Colores)
        fill_header_gray = PatternFill(start_color='E7E6E6', end_color='E7E6E6',
                                       fill_type='solid')  # Gris claro para encabezados inferiores
        fill_header_blue = PatternFill(start_color='DAE8FC', end_color='DAE8FC',
                                       fill_type='solid')  # Azul claro (Roles, Competencias)

        # Alineación
        center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 4. Construcción del Encabezado (Filas 1 y 2)
        # Título Principal
        ws.merge_cells('D1:J1')
        ws['D1'] = "FORMATO DE VALORACION DE PUESTOS"
        ws['D1'].font = font_header_main
        ws['D1'].alignment = center_aligned

        # Subtítulo (Dinámico o estático según requerimiento)
        ws.merge_cells('D2:J2')
        ws['D2'] = "PUESTOS INSTITUCIONALES LEGALIZADOS"
        ws['D2'].font = font_header_main
        ws['D2'].alignment = center_aligned

        # 5. Encabezados de Tabla (Filas 3 y 4)
        headings_l1 = [
            ("NRO.", "A3:A4"),
            ("DENOMINACION DE PUESTO", "B3:B4"),
            ("RESPONSABILIDAD", "C3:C3"),  # Padre de Rol
            ("COMPETENCIAS", "D3:E3"),  # Padre de Instruccion y Experiencia
            ("NIVEL DE COMPLEJIDAD DEL PUESTO", "F3:H3"),  # Padre de Decisiones, Impacto, Complejidad
            ("CLASIFICACION", "I3:I3"),  # Padre de Grupo
            ("REFERENCIAL", "J3:J4"),  # Empleado (Columna extra solicitada)
        ]

        headings_l2 = [
            # Col C
            ("ROL DEL PUESTO", "C4"),
            # Cols D-E
            ("INSTRUCCIÓN FORMAL", "D4"),
            ("EXPERIENCIA", "E4"),
            # Cols F-H
            ("TOMA DE DECISIONES", "F4"),
            ("IMPACTO INSTITUCIONAL A RESULTADOS", "G4"),
            ("COMPLEJIDAD TECNICA (ESPECIALIZACION DE TAREAS)", "H4"),
            # Col I
            ("GRUPO OCUPACIONAL", "I4"),
        ]

        # Aplicar Merge y Texto Nivel 1
        for text, merge_range in headings_l1:
            ws.merge_cells(merge_range)
            # Obtenemos la celda superior izquierda del merge
            top_left_cell = ws[merge_range.split(':')[0]]
            top_left_cell.value = text
            top_left_cell.font = font_header_sub
            top_left_cell.alignment = center_aligned
            top_left_cell.border = thin_border
            top_left_cell.fill = fill_header_blue

        # Aplicar Texto Nivel 2
        for text, cell_coord in headings_l2:
            cell = ws[cell_coord]
            cell.value = text
            cell.font = Font(name='Arial', size=8, bold=True)
            cell.alignment = center_aligned
            cell.border = thin_border
            cell.fill = fill_header_blue

        # 6. Datos (Sólo Legalizados: tienen las 3 firmas)
        # Filtro: prepared_by, reviewed_by, approved_by no nulos
        profiles = JobProfile.objects.select_related(
            'job_role', 'required_instruction', 'decision_making',
            'management_impact', 'final_complexity_level',
            'occupational_classification', 'referential_employee__person'
        ).filter(
            is_active=True,
            prepared_by__isnull=False,
            reviewed_by__isnull=False,
            approved_by__isnull=False
        ).order_by('occupational_classification__grade', 'specific_job_title')

        row_num = 5
        for idx, p in enumerate(profiles, start=1):
            # Obtener textos seguros
            job_title = p.specific_job_title
            role = p.job_role.name if p.job_role else ""
            instruction = p.required_instruction.name if p.required_instruction else ""
            experience = f"{p.required_experience_months} Meses"
            decision = p.decision_making.name if p.decision_making else ""
            impact = p.management_impact.name if p.management_impact else ""
            complexity = p.final_complexity_level.name if p.final_complexity_level else ""
            group = p.occupational_classification.occupational_group if p.occupational_classification else ""

            ref_employee = str(p.referential_employee.person) if p.referential_employee else "VACANTE"

            row = [
                idx,  # A
                job_title,  # B
                role,  # C
                instruction,  # D
                experience,  # E
                decision,  # F
                impact,  # G
                complexity,  # H
                group,  # I
                ref_employee  # J
            ]

            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = font_body
                cell.border = thin_border
                # Alinear al centro excepto Titulo y Referencial que pueden ser largos
                if col_idx in [2, 10]:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = center_aligned

            row_num += 1

        # 7. Ajuste de Ancho de Columnas
        column_widths = [5, 40, 25, 25, 15, 20, 20, 20, 25, 35]  # A to J
        for i, width in enumerate(column_widths, start=1):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # Guardar
        wb.save(response)
        return response


def api_get_profile_valuation_chain(request, profile_id):
    """
    Devuelve los 6 niveles de valoración con los valores seleccionados del perfil.
    Devuelve los catalog_item_id seleccionados para cada nivel.
    """
    try:
        profile = JobProfile.objects.filter(pk=profile_id).select_related(
            'job_role',
            'required_instruction',
            'decision_making',
            'management_impact',
            'final_complexity_level'
        ).first()

        if not profile:
            return JsonResponse({
                'success': False,
                'message': 'Perfil no encontrado'
            }, status=404)

        # Construir respuesta con los catalog_items seleccionados
        valuation_data = {
            'success': True,
            'selected_values': {
                'role_id': profile.job_role_id,
                'instruction_id': profile.required_instruction_id,
                'experience_months': profile.required_experience_months or 0,
                'decision_id': profile.decision_making_id,
                'impact_id': profile.management_impact_id,
                'complexity_id': profile.final_complexity_level_id
            }
        }

        return JsonResponse(valuation_data)

    except Exception as e:
        import traceback
        print("=" * 80)
        print("ERROR en api_get_profile_valuation_chain:")
        traceback.print_exc()
        print("=" * 80)
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)


class AssignGroupApiView(LoginRequiredMixin, View):
    """
    API para asignar manualmente el grupo ocupacional (OccupationalMatrix) a un perfil.
    Recibe JSON: { profile_id: <int>, matrix_id: <int> }
    """

    def post(self, request):
        try:
            data = json.loads(request.body)
            profile_id = data.get('profile_id')
            matrix_id = data.get('matrix_id')

            if not profile_id or not matrix_id:
                return JsonResponse({'success': False, 'message': 'Faltan datos requeridos (profile_id, matrix_id)'},
                                    status=400)

            profile = get_object_or_404(JobProfile, pk=profile_id)
            classification = get_object_or_404(OccupationalMatrix, pk=matrix_id)

            # Validar si es necesario alguna regla de negocio aquí.
            # Por ahora asignamos directo.
            profile.occupational_classification = classification

            # Copiar atributos de la Matriz al Perfil para mantener consistencia
            profile.job_role = classification.required_role
            profile.required_instruction = classification.minimum_instruction
            profile.decision_making = classification.required_decision
            profile.management_impact = classification.required_impact
            profile.final_complexity_level = classification.complexity_level
            profile.required_experience_months = classification.minimum_experience_months

            profile.save()

            return JsonResponse({
                'success': True,
                'message': f'Perfil asignado correctamente a {classification.occupational_group} - Grado {classification.grade}'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)


class JobActivityReportExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Genera el reporte de Levantamiento de Actividades (Formato MDT)
    Calcula dinámicamente los valores AG, F, C y Total.
    """
    permission_required = "function_manual.view_jobprofile"

    def get(self, request, pk):
        profile = get_object_or_404(JobProfile.objects.select_related('administrative_unit'), pk=pk)
        activities = profile.activities.all().select_related(
            'action_verb', 'deliverable', 'complexity', 'contribution', 'frequency'
        )

        # 1. Configuración del Libro y Hoja
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Levantamiento_Actividades_{profile.position_code}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Levantamiento de Actividades"

        # 2. Estilos Globales
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='DAE8FC', end_color='DAE8FC', fill_type='solid')
        main_header_font = Font(name='Arial', size=11, bold=True)
        sub_header_font = Font(name='Arial', size=8, bold=True)
        body_font = Font(name='Arial', size=9)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 3. Encabezado Superior (Fila 1 y 2)
        ws.merge_cells('A1:C2')
        ws['A1'] = "Municipio de Loja"
        ws['A1'].font = main_header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('D1:K2')
        ws['D1'] = "FORMATO LEVANTAMIENTO DE ACTIVIDADES"
        ws['D1'].font = main_header_font
        ws['D1'].alignment = center_align

        # Denominación del Puesto (Fila 3)
        ws.merge_cells('A3:C3')
        ws['A3'] = "DENOMINACIÓN DE PUESTO:"
        ws['A3'].font = sub_header_font
        ws.merge_cells('D3:K3')
        ws['D3'] = f"{profile.specific_job_title} ({profile.position_code})"
        ws['D3'].font = body_font

        # 4. Cabeceras de Tabla (Fila 4 y 5)
        headers = [
            ("NRO.", "A4:A5"),
            ("ENTREGABLES\n(PRODUCTOS/SERVICIOS)", "B4:B5"),
            ("ACTIVIDADES", "C4:C5"),
            ("APORTE A LA GESTION\n(AG)", "D4:D5"),
            ("FRECUENCIA (F)", "E4:E5"),
            ("COMPLEJIDAD (C)", "F4:F5"),
            ("AG", "G4:G5"),
            ("F", "H4:H5"),
            ("C", "I4:I5"),
            ("TOTAL\nAG*(F+C)", "J4:J5"),
            ("ACTIVIDADES\nSELECCIONADAS PARA EL\nPERFIL", "K4:K5"),
        ]

        for text, merge_range in headers:
            ws.merge_cells(merge_range)
            cell = ws[merge_range.split(':')[0]]
            cell.value = text
            cell.font = sub_header_font
            cell.alignment = center_align
            cell.fill = header_fill
            # Aplicar bordes al rango mergeado
            for row in ws[merge_range]:
                for c in row:
                    c.border = thin_border

        # 5. Mapeos de Lógica Normativa
        def get_val_ag_c(name):
            n = name.upper()
            if "ALTO" in n: return 3
            if "MEDIO" in n: return 2
            if "BAJO" in n: return 1
            return 0

        def get_val_f(name):
            n = name.upper()
            if "DIARIO" in n: return 5
            if "SEMANAL" in n: return 4
            if "MENSUAL" in n: return 3
            if "TRIMESTRAL" in n or "SEMESTRAL" in n: return 2
            if "ANUAL" in n: return 1
            return 0

        # 6. Datos de Actividades
        row_num = 6
        for idx, act in enumerate(activities, start=1):
            # Obtención de nombres
            deliverable_name = act.deliverable.name if act.deliverable else "N/A"
            activity_text = f"{act.action_verb.name} {act.description}".strip()
            ag_name = act.contribution.name if act.contribution else "Bajo"
            f_name = act.frequency.name if act.frequency else "Anual"
            c_name = act.complexity.name if act.complexity else "Bajo"

            # Obtención de valores numéricos
            v_ag = get_val_ag_c(ag_name)
            v_f = get_val_f(f_name)
            v_c = get_val_ag_c(c_name)
            total = v_ag * (v_f + v_c)

            data_row = [
                idx,  # A: Nro
                deliverable_name,  # B: Entregable
                activity_text,  # C: Actividad (Verbo + Desc)
                ag_name,  # D: Nombre Aporte
                f_name,  # E: Nombre Frecuencia
                c_name,  # F: Nombre Complejidad
                v_ag,  # G: Valor AG
                v_f,  # H: Valor F
                v_c,  # I: Valor C
                total,  # J: Total
                "SELECCIONADA"  # K: Selección
            ]

            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = body_font
                cell.border = thin_border
                # Alineación
                if col_idx in [2, 3]:  # Texto largo a la izquierda
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

            row_num += 1

        # 7. Ajuste de Ancho de Columnas
        widths = [5, 25, 45, 12, 12, 12, 5, 5, 5, 10, 15]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(response)
        return response
